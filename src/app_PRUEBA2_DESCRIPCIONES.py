# ================== IMPORTS ==================
# Importa librerías esenciales para el funcionamiento del script.
# asyncio: para operaciones de E/S concurrentes (lecturas OPC UA).
# threading: para ejecutar el dashboard y el monitoreo en hilos separados.
# datetime, timedelta, time: para manejar fechas y duraciones.
# os, uuid, Path, sys: para la gestión del sistema de archivos, variables de entorno y rutas.
# logging, RotatingFileHandler: para registrar eventos y errores en un archivo rotativo.
# db (oracledb/cx_Oracle): para la conexión con la base de datos Oracle.
# asyncua: el cliente para comunicarse con el servidor OPC UA.
# dash, dcc, html, dash_table: para construir el dashboard web.
# openpyxl: para leer las descripciones de los archivos Excel.
# requests: para enviar notificaciones a Telegram.
# html: para escapar caracteres especiales en los mensajes de Telegram.
# flask: para la respuesta HTTP del servidor web de Dash.
# BadNodeIdUnknown: excepción específica de asyncua para nodos no encontrados.
# dotenv: para cargar variables de entorno desde un archivo .env.

import asyncio
import threading
from datetime import datetime, timedelta, time
import os
import uuid
from typing import Optional
from pathlib import Path
import sys
import logging
from logging.handlers import RotatingFileHandler

# DB: preferimos oracledb (thin) o cx_Oracle si está
try:
    import cx_Oracle as db
except ModuleNotFoundError:
    import oracledb as db

from asyncua import Client, ua
from dash import Dash, dcc, html
from dash.dependencies import Input, Output
from openpyxl import load_workbook
from dash import dash_table
import requests
import html as html_escape 
from flask import Response 

# Excepciones OPC UA (según versión)
try:
    from asyncua.ua.uaerrors import BadNodeIdUnknown
except Exception:
    try:
        from asyncua.ua.uaerrors._auto import BadNodeIdUnknown  # fallback
    except Exception:
        class BadNodeIdUnknown(Exception):
            pass

# (Opcional) Cargar .env si existiera (una sola vez)
try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover
    load_dotenv = None

# ================== RUTAS Y ENTORNO ==================
# Define la ruta base del script, lo que permite que sea portable.
# Si el script está empaquetado (por ejemplo, con PyInstaller), _MEIPASS es la ruta temporal.
BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
if load_dotenv:
    load_dotenv(BASE_DIR / ".env")

# -------------------- Archivos Excel para descripción (ÚNICO lugar) --------------------
# Diccionario que mapea los grupos de nodos a sus archivos de descripciones en Excel.
ARCHIVOS_DESCRIPCIONES = {
    "BOMBAS_SH": str((BASE_DIR / "GMFs_de_Equipos" / "GMF_BOMBAS_SH.xlsx").resolve()),
    "CONVEYOR_SH": str((BASE_DIR / "GMFs_de_Equipos" / "GMF_CONVEYOR_SH.xlsx").resolve()),
}

# ================== CONFIGURACIÓN GENERAL ==================
# Carga las configuraciones desde las variables de entorno.
# Permite cambiar fácilmente la URL del servidor OPC UA, los grupos de nodos,
# el rango de GMFs y el intervalo de lectura sin modificar el código.
URL = os.getenv("OPCUA_URL", "opc.tcp://localhost:52250/freeopcua/server/")
NODE_GROUPS = os.getenv("NODE_GROUPS", "BOMBAS_SH,CONVEYOR_SH").split(",")
NODE_GROUPS = [g.strip() for g in NODE_GROUPS if g.strip()]
START_HEX = int(os.getenv("START_HEX", "0x80"), 16)
END_HEX = int(os.getenv("END_HEX", "0xFF"), 16)
INTERVALO_SEGUNDOS = float(os.getenv("INTERVALO_SEGUNDOS", "0.02"))

# ================== TELEGRAM ==================
# Configuración específica para las notificaciones de Telegram.
# TELEGRAM_TOKEN: el token del bot.
# TELEGRAM_CHAT_IDS: la lista de IDs de chat a los que se enviarán los mensajes.
# TELEGRAM_THRESHOLD_SECONDS: el umbral de tiempo para enviar la primera alerta.
TELEGRAM_ENABLED = os.getenv("TELEGRAM_ENABLED", "true").lower() == "true"
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN") 
CHAT_IDS_ENV = os.getenv("TELEGRAM_CHAT_IDS")
if CHAT_IDS_ENV:
    TELEGRAM_CHAT_IDS = [x.strip() for x in CHAT_IDS_ENV.split(",") if x.strip()]
else:
    TELEGRAM_CHAT_IDS = [
        "1600754452",
        "5015132163",
    ]

TELEGRAM_THRESHOLD_SECONDS = int(os.getenv("TELEGRAM_THRESHOLD_SECONDS", "60"))
TELEGRAM_RESTABLECIDO_UNICO = os.getenv("TELEGRAM_RESTABLECIDO_UNICO", "true").lower() == "true"

# ================== LOGGING ==================
# Configura el sistema de logging para registrar eventos, advertencias y errores.
# Los logs se escriben en un archivo rotativo (`monitor.log`) para evitar que crezca indefinidamente.
# También se imprimen en la consola (`sys.stdout`).
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
LOG_FILE = os.getenv("LOG_FILE", str((BASE_DIR / "monitor.log").resolve()))
logger = logging.getLogger("opcua-monitor")
logger.setLevel(LOG_LEVEL)

fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s", "%Y-%m-%d %H:%M:%S")
fh = RotatingFileHandler(LOG_FILE, maxBytes=1_000_000, backupCount=3, encoding="utf-8")
fh.setFormatter(fmt)
sh = logging.StreamHandler(sys.stdout)
sh.setFormatter(fmt)
if not logger.handlers:
    logger.addHandler(fh)
    logger.addHandler(sh)
else:
    logger.handlers.clear()
    logger.addHandler(fh)
    logger.addHandler(sh)

# ================== ESTADO ==================
# Variables globales para mantener el estado del monitoreo y la interfaz web.
# Se usa un lock de threading para evitar condiciones de carrera al acceder
# a estas variables desde múltiples hilos (monitoreo y dashboard).
lock = threading.Lock()

tiempos_gmf = {}            # { "GRUPO.GMFxxxx": datetime_inicio } - Guarda el inicio de cada falla.
tiempos_congelados = {}     # { "GRUPO.GMFxxxx": seg_vistos } - Almacena el tiempo de falla cuando el equipo no está en 'continuous'.
acumulado_continuous = {}   # { "GRUPO": timedelta_off } - Acumula el tiempo de falla solo cuando 'continuous' está en 0.
ultima_marca_continuous = {} # { "GRUPO": datetime_ultima_marca } - Marca la última vez que se leyó 'continuous' para calcular el tiempo acumulado.
descripciones_gmf = {}      # Diccionario para almacenar las descripciones de las GMFs desde el Excel.

# Telegram: flags para controlar el envío de mensajes y evitar spam.
tg_enviado = {g: False for g in NODE_GROUPS}
tg_restaurado_enviado = {g: False for g in NODE_GROUPS}
last_continuous_state = {g: None for g in NODE_GROUPS}

# Monitoreo Oracle
SESSION_ID = uuid.uuid4().hex
last_mon_update = datetime.min

# Error de conexión (global)
error_activo = False
error_inicio = None
ultimo_error_msg = ""

# Error por grupo (ventanas inicio-fin)
error_grupo_activo = {g: False for g in NODE_GROUPS}
error_grupo_inicio = {g: None for g in NODE_GROUPS}
error_grupo_msg = {g: "" for g in NODE_GROUPS}

# ================== TURNOS ==================
def determinar_turno(fecha: datetime) -> str:
    """Calcula el turno de trabajo basado en la hora de una fecha dada."""
    hora = fecha.time()
    if time(6, 5) <= hora <= time(14, 30):
        return "MAÑANA"
    elif time(14, 39) <= hora <= time(22, 54):
        return "TARDE"
    elif time(23, 13) <= hora <= time(23, 59) or time(0, 0) <= hora <= time(5, 55):
        return "NOCHE"
    return "FUERA DE TURNO"

# ================== ORACLE ==================
# Funciones para manejar la conexión y la persistencia de datos en la base de datos Oracle.
# Estas funciones registran el inicio y fin del monitoreo, así como los errores de conexión.
def _oracle_conn():
    """Establece y devuelve una conexión a la base de datos Oracle."""
    user = os.getenv("ORACLE_USER")
    pwd  = os.getenv("ORACLE_PWD")
    dsn  = os.getenv("ORACLE_DSN")
    if not all([user, pwd, dsn]):
        raise RuntimeError("Faltan ORACLE_USER/ORACLE_PWD/ORACLE_DSN en .env")
    return db.connect(user=user, password=pwd, dsn=dsn)

def _ensure_tablas_oracle():
    """Verifica si las tablas de la base de datos existen y las crea si no."""
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        cur.execute(
            """
            BEGIN
                EXECUTE IMMEDIATE '
                    CREATE TABLE MONITOREO_OPCUA (
                        SESSION_ID      VARCHAR2(64),
                        TIPO            VARCHAR2(30),
                        GRUPO           VARCHAR2(50),
                        INICIO          TIMESTAMP,
                        FIN             TIMESTAMP,
                        MENSAJE         VARCHAR2(4000)
                    )';
            EXCEPTION
                WHEN OTHERS THEN
                    IF SQLCODE != -955 THEN RAISE; END IF;
            END;
            """
        )
        conn.commit()
    except Exception as e:
        logger.error(f"Error asegurando tablas Oracle: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def iniciar_monitoreo_en_oracle(inicio: datetime):
    """Registra una nueva sesión de monitoreo en la tabla MONITOREO_OPCUA."""
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        _ensure_tablas_oracle()
        cur.execute(
            """
            INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
            VALUES (:sid, 'MONITOREO', NULL, :ini, :fin, :msg)
            """,
            dict(sid=SESSION_ID, ini=inicio, fin=inicio, msg=None)
        )
        conn.commit()
        logger.info(f"Sesión de monitoreo iniciada: {inicio} | SESSION_ID={SESSION_ID}")
    except Exception as e:
        logger.error(f"Error iniciando monitoreo en Oracle: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def actualizar_fin_monitoreo(fin: datetime):
    """Actualiza la fecha y hora de fin de la sesión de monitoreo en curso."""
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        cur.execute(
            """
            UPDATE MONITOREO_OPCUA
            SET FIN = :fin
            WHERE SESSION_ID = :sid AND TIPO = 'MONITOREO'
            """,
            dict(fin=fin, sid=SESSION_ID)
        )
        conn.commit()
    except Exception as e:
        logger.error(f"Error actualizando fin de monitoreo: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def registrar_error_conexion(inicio: datetime, fin: datetime, mensaje: str):
    """Registra una ventana de tiempo en la que la conexión OPC UA falló."""
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        _ensure_tablas_oracle()
        cur.execute(
            """
            INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
            VALUES (:sid, 'ERROR_CONEXION', NULL, :ini, :fin, :msg)
            """,
            dict(sid=SESSION_ID, ini=inicio, fin=fin, msg=(mensaje or "")[:3999])
        )
        conn.commit()
        logger.warning(f"Error de conexión registrado: {inicio} -> {fin} | {mensaje}")
    except Exception as e:
        logger.error(f"Error registrando error de conexión: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def abrir_error_grupo(grupo: str, inicio: datetime, mensaje: str):
    """Registra el inicio de un período de error para un grupo específico de nodos."""
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        _ensure_tablas_oracle()
        cur.execute(
            """
            INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
            VALUES (:sid, 'ERROR_GRUPO', :grp, :ini, NULL, :msg)
            """,
            dict(sid=SESSION_ID, grp=grupo, ini=inicio, msg=(mensaje or "")[:3999])
        )
        conn.commit()
        logger.warning(f"ERROR_GRUPO abierto: {grupo} @ {inicio} | {mensaje}")
    except Exception as e:
        logger.error(f"Error abriendo ERROR_GRUPO ({grupo}): {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def cerrar_error_grupo(grupo: str, fin: datetime):
    """Actualiza la fecha y hora de fin de una ventana de error de grupo que estaba abierta."""
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        cur.execute(
            """
            UPDATE MONITOREO_OPCUA
            SET FIN = :fin
            WHERE SESSION_ID = :sid
              AND TIPO = 'ERROR_GRUPO'
              AND GRUPO = :grp
              AND FIN IS NULL
            """,
            dict(fin=fin, sid=SESSION_ID, grp=grupo)
        )
        conn.commit()
        logger.info(f"ERROR_GRUPO cerrado: {grupo} @ {fin}")
    except Exception as e:
        logger.error(f"Error cerrando ERROR_GRUPO ({grupo}): {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

# ================== HELPERS DESCRIPCIONES ==================
# Funciones auxiliares para cargar y buscar descripciones de GMFs.
def _normalize_code(codigo) -> str:
    """Normaliza un código GMF a formato de string para la búsqueda."""
    if codigo is None:
        return ""
    if isinstance(codigo, (int, float)):
        s = f"{codigo}"
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s.strip().upper()
    return str(codigo).strip().upper()

def _desc_gmf(grupo: str, gmf_codigo: str) -> str:
    """Busca y devuelve la descripción de un código GMF. Si no la encuentra, devuelve una cadena por defecto."""
    return descripciones_gmf.get(grupo, {}).get(gmf_codigo.strip().upper(), "Sin descripción")

def _construir_mensaje_telegram(grupo: str, segundos: int, activos: list[str]) -> str:
    """Crea el cuerpo del mensaje de alerta de Telegram con formato HTML."""
    if not activos:
        return f'⚠️ {html_escape.escape(grupo)} en falla hace {segundos} s.'
    if len(activos) == 1:
        g = activos[0]
        d = _desc_gmf(grupo, g)
        return f'⚠️ {html_escape.escape(grupo)} con falla {html_escape.escape(g)}: {html_escape.escape(d)} hace {segundos} s.'
    partes = [f'{html_escape.escape(g)}: {html_escape.escape(_desc_gmf(grupo, g))}' for g in activos]
    listado = "; ".join(partes[:10])
    extra = "" if len(partes) <= 10 else f" (+{len(partes)-10} más)"
    return f'⚠️ {html_escape.escape(grupo)} con fallas <b>{listado}{extra}</b> hace {segundos} s.'

# ================== TELEGRAM ==================
# Funciones para la comunicación con la API de Telegram.
def _telegram_config_ok() -> bool:
    """Verifica que la configuración de Telegram sea válida."""
    if not TELEGRAM_ENABLED:
        logger.info("Telegram desactivado por configuración")
        return False
    if not TELEGRAM_TOKEN:
        logger.warning("Falta TELEGRAM_TOKEN. Deshabilitando envíos Telegram.")
        return False
    if not TELEGRAM_CHAT_IDS or not any(x.strip() and x.strip().lstrip("-").isdigit() for x in TELEGRAM_CHAT_IDS):
        logger.warning("TELEGRAM_CHAT_IDS vacío o inválido. Deshabilitando Telegram.")
        return False
    return True

def enviar_telegram_a_todos(body: str):
    """Envía un mensaje a todos los chats configurados en TELEGRAM_CHAT_IDS."""
    if not _telegram_config_ok():
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    sess = requests.Session()
    for cid in TELEGRAM_CHAT_IDS:
        cid = cid.strip()
        if not cid or not cid.lstrip("-").isdigit():
            logger.warning(f"chat_id inválido en lista: {cid!r}. Omitiendo.")
            continue
        payload = {
            "chat_id": cid,
            "text": body,
            "parse_mode": "HTML",
            "disable_web_page_preview": True,
        }
        try:
            r = sess.post(url, json=payload, timeout=(3, 8))
            if r.status_code == 200:
                logger.info(f"Telegram enviado a chat_id={cid}")
            else:
                logger.error(f"Error enviando a chat_id={cid}: {r.status_code} {r.text}")
        except Exception as e:
            logger.error(f"Excepción enviando a chat_id={cid}: {e}")

def enviar_telegram_restaurado(grupo: str, segundos: int):
    """Envía un mensaje de restauración (equipo arreglado) a Telegram."""
    if not _telegram_config_ok():
        return
    grupo_txt = html_escape.escape(grupo)
    cuerpo = f'✅ Equipo "{grupo_txt}" restablecido. Tiempo acumulado {int(max(0, segundos))} s'
    enviar_telegram_a_todos(cuerpo)

# ================== CARGA DE DESCRIPCIONES ==================
def cargar_descripciones():
    """Lee los archivos Excel y carga las descripciones de las GMFs en un diccionario global."""
    for grupo, ruta in ARCHIVOS_DESCRIPCIONES.items():
        descripciones_gmf[grupo] = {}
        try:
            wb = load_workbook(ruta, data_only=True)
            hoja = wb.active
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                codigo_raw, descripcion = fila[:2]
                if codigo_raw and descripcion:
                    codigo = _normalize_code(codigo_raw)
                    if not codigo.startswith("GMF"):
                        codigo = "GMF" + codigo
                    descripciones_gmf[grupo][codigo] = str(descripcion).strip()
            logger.info(f"[Desc] {grupo}: {len(descripciones_gmf[grupo])} códigos cargados desde {ruta}")
        except Exception as e:
            logger.error(f"Error al cargar descripciones de {grupo} desde {ruta}: {e}")

# ================== EXPORTAR A ORACLE (historial de fallas) ==================
def exportar_a_oracle(grupo, gmf, descripcion, inicio, fin, acumulado):
    """
    Exporta el registro de una falla completa a una tabla específica en Oracle.
    Crea la tabla si no existe y luego inserta los datos de la falla.
    """
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        tabla = f"HISTORIAL_FALLAS_{grupo.upper()}"
        turno = determinar_turno(inicio)
        cur.execute(
            f"""
            BEGIN
                EXECUTE IMMEDIATE '
                    CREATE TABLE {tabla} (
                        EQUIPO          VARCHAR2(50),
                        FALLA           VARCHAR2(50),
                        DESCRIPCION     VARCHAR2(255),
                        INICIO_FALLA    TIMESTAMP,
                        FIN_FALLA       TIMESTAMP,
                        TIEMPO_OFF_SEG  NUMBER,
                        TURNO           VARCHAR2(20)
                    )';
            EXCEPTION WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
            END;
            """
        )
        cur.execute(
            f"""
            INSERT INTO {tabla} (EQUIPO, FALLA, DESCRIPCION, INICIO_FALLA, FIN_FALLA, TIEMPO_OFF_SEG, TURNO)
            VALUES (:1, :2, :3, :4, :5, :6, :7)
            """,
            (grupo, gmf, descripcion, inicio, fin, acumulado, turno)
        )
        conn.commit()
        logger.info(f"Exportado a Oracle: {grupo} | {gmf} | {turno} | {acumulado} seg")
    except Exception as e:
        logger.error(f"Error al exportar a Oracle: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

# ================== NODE CACHE ==================
class NodeCache:
    """
    Clase para crear y almacenar una caché de los nodos OPC UA por grupo.
    Esto evita tener que buscar los nodos en el servidor en cada ciclo de lectura,
    mejorando la eficiencia y el rendimiento.
    """
    def __init__(self, client: Client):
        self.client = client
        self.cache: dict[str, dict] = {}

    def build_for_groups(self, grupos: list[str]):
        """Construye la caché de nodos para los grupos especificados."""
        for grupo in grupos:
            total_fault = self.client.get_node(f"ns=2;s={grupo}.TOTAL_FAULT")
            continuous  = self.client.get_node(f"ns=2;s={grupo}.CONTINUOUS")
            total_start = self.client.get_node(f"ns=2;s={grupo}.TOTAL_START")

            hex_suffixes = [f"{i:X}" for i in range(START_HEX, END_HEX + 1)]
            node_ids = [ua.NodeId.from_string(f"ns=2;s={grupo}.GMF{suf}") for suf in hex_suffixes]
            gmf_nodes = [self.client.get_node(nid) for nid in node_ids]

            self.cache[grupo] = {
                "total_fault": total_fault,
                "continuous": continuous,
                "total_start": total_start,
                "gmf_nodes": gmf_nodes,
                "gmf_suffixes": hex_suffixes,
                "gmf_validated": False,
            }
        return self

    def for_group(self, grupo: str) -> dict:
        """Devuelve los nodos en caché para un grupo específico."""
        return self.cache[grupo]

# ================== LECTURA POR GRUPO ==================
async def _validate_gmfs(client: Client, grupo: str, ncache: dict):
    """Valida la existencia de cada nodo GMF en el servidor para evitar errores de lectura."""
    valid_nodes = []
    valid_suffixes = []
    for node, suf in zip(ncache["gmf_nodes"], ncache["gmf_suffixes"]):
        try:
            await node.read_value()
            valid_nodes.append(node)
            valid_suffixes.append(suf)
        except BadNodeIdUnknown:
            logger.debug(f'{grupo}: GMF{suf} no existe, se omite.')
        except Exception:
            logger.debug(f'{grupo}: GMF{suf} no legible por ahora, se omite.')
    ncache["gmf_nodes"] = valid_nodes
    ncache["gmf_suffixes"] = valid_suffixes
    ncache["gmf_validated"] = True

async def leer_grupo_si_falla(client: Client, grupo: str, ncache: dict):
    """
    Función principal para leer los nodos OPC UA de un grupo.
    Si el nodo `TOTAL_FAULT` está activo, lee los nodos GMF.
    Maneja el estado `CONTINUOUS` para exportar datos y enviar notificaciones.
    """
    ahora = datetime.now()
    try:
        # --- TOTAL_START opcional ---
        # Verifica si el equipo está "arrancado" para monitorear sus fallas.
        started = True
        ts_node = ncache.get("total_start")
        if ts_node is not None:
            try:
                v = await ts_node.read_value()
                started = bool(v)
            except BadNodeIdUnknown:
                logger.warning(f'{grupo}: nodo TOTAL_START no existe. Se asume iniciado.')
                ncache["total_start"] = None
            except Exception as e:
                logger.warning(f'{grupo}: error leyendo TOTAL_START ({e}). Se asume iniciado.')

        if not started:
            await asyncio.sleep(0.5)
            return

        nodo_fault = ncache["total_fault"]
        nodo_cont  = ncache["continuous"]
        valor_fault, valor_continuous = await client.read_values([nodo_fault, nodo_cont])

        # Transición CONTINUOUS
        # Si el valor de `CONTINUOUS` cambia de True a False, resetea el flag de restauración de Telegram.
        prev = last_continuous_state.get(grupo)
        last_continuous_state[grupo] = bool(valor_continuous)
        if prev is True and valor_continuous is False:
            tg_restaurado_enviado[grupo] = False

        # Acumulador CONTINUOUS=0
        # Acumula el tiempo de falla si el equipo NO está en modo 'continuous'.
        with lock:
            if grupo not in acumulado_continuous:
                acumulado_continuous[grupo] = timedelta(0)
                ultima_marca_continuous[grupo] = ahora
            if not valor_continuous:
                delta = ahora - ultima_marca_continuous[grupo]
                acumulado_continuous[grupo] += delta
            ultima_marca_continuous[grupo] = ahora

            segundos_acum = int(acumulado_continuous[grupo].total_seconds())
            activos_snapshot = [gmf_id.split(".")[1]
                                 for gmf_id in tiempos_gmf.keys()
                                 if gmf_id.startswith(grupo + ".")]
            debe_enviar = (segundos_acum >= TELEGRAM_THRESHOLD_SECONDS) and (not tg_enviado.get(grupo, False))
            if debe_enviar:
                tg_enviado[grupo] = True

        if debe_enviar:
            body = _construir_mensaje_telegram(segundos_acum, activos_snapshot)
            asyncio.create_task(asyncio.to_thread(enviar_telegram_a_todos, body))

        # Lectura de GMFs
        try:
            values = await client.read_values(ncache["gmf_nodes"])
        except BadNodeIdUnknown:
            if not ncache.get("gmf_validated", False):
                await _validate_gmfs(client, grupo, ncache)
                if not ncache["gmf_nodes"]:
                    with lock:
                        estaba_activo = error_grupo_activo.get(grupo)
                    if estaba_activo:
                        cerrar_error_grupo(grupo, datetime.now())
                        with lock:
                            error_grupo_activo[grupo] = False
                            error_grupo_inicio[grupo] = None
                            error_grupo_msg[grupo] = ""
                    return
                values = await client.read_values(ncache["gmf_nodes"])
            else:
                raise

        gmfs_detectadas = set()

        for suffix, value in zip(ncache["gmf_suffixes"], values):
            if isinstance(value, int):
                for bit_pos in range(16):
                    if (value >> bit_pos) & 1:
                        gmf_base = f"{suffix}{bit_pos:X}"
                        gmf_id = f"{grupo}.GMF{gmf_base.upper()}"
                        with lock:
                            gmfs_detectadas.add(gmf_id)
                            if gmf_id not in tiempos_gmf:
                                tiempos_gmf[gmf_id] = ahora
                            segundos = int((ahora - tiempos_gmf[gmf_id]).total_seconds())
                            tiempos_congelados[gmf_id] = segundos

        # GMFs que se inactivaron
        with lock:
            inactivas = [gmf for gmf in list(tiempos_gmf.keys())
                         if gmf.startswith(grupo) and gmf not in gmfs_detectadas]

        for gmf in inactivas:
            grupo_actual, gmf_codigo = gmf.split(".")
            val_fault, val_cont = await client.read_values([nodo_fault, nodo_cont])
            descripcion = _desc_gmf(grupo, gmf_codigo)

            with lock:
                inicio_falla = tiempos_gmf.get(gmf)
                acumulado = int(acumulado_continuous[grupo].total_seconds())

            if bool(val_cont):
                # Exporta y limpia (se terminó y el equipo está operativo)
                exportar_a_oracle(grupo, gmf_codigo, descripcion, inicio_falla, ahora, acumulado)

                debe_enviar_rest = True
                if TELEGRAM_RESTABLECIDO_UNICO and tg_restaurado_enviado.get(grupo, False):
                    debe_enviar_rest = False

                if debe_enviar_rest and (acumulado > TELEGRAM_THRESHOLD_SECONDS):
                    asyncio.create_task(asyncio.to_thread(enviar_telegram_restaurado, grupo, acumulado))
                    if TELEGRAM_RESTABLECIDO_UNICO:
                        tg_restaurado_enviado[grupo] = True

                with lock:
                    tiempos_gmf.pop(gmf, None)
                    tiempos_congelados.pop(gmf, None)
            else:
                # CONTINUOUS=0 -> mantenemos visible con el contador congelado
                with lock:
                    pass

        # Reset acumulador/flags si no hay GMF activas y CONTINUOUS=1
        with lock:
            if (valor_continuous is True) and (not any(k.startswith(grupo) for k in tiempos_gmf)):
                acumulado_continuous[grupo] = timedelta(0)
                tg_enviado[grupo] = False

        # Cierre de ventana de error por grupo si estaba abierta
        with lock:
            estaba_activo = error_grupo_activo.get(grupo)
        if estaba_activo:
            cerrar_error_grupo(grupo, datetime.now())
            with lock:
                error_grupo_activo[grupo] = False
                error_grupo_inicio[grupo] = None
                error_grupo_msg[grupo] = ""

    except Exception as e:
        with lock:
            ya_activo = error_grupo_activo.get(grupo)
        if not ya_activo:
            inicio = datetime.now()
            msg = f"Error en grupo {grupo}: {e}"
            logger.error(msg)
            abrir_error_grupo(grupo, inicio, msg)
            with lock:
                error_grupo_activo[grupo] = True
                error_grupo_inicio[grupo] = inicio
                error_grupo_msg[grupo] = msg

# ================== LOOP PRINCIPAL ==================
async def main():
    """
    Bucle principal asíncrono. Se encarga de la conexión con el servidor OPC UA
    y la lectura continua de los nodos.
    Maneja la reconexión en caso de fallas y registra los errores en Oracle.
    """
    cargar_descripciones()
    try:
        _ensure_tablas_oracle()
    except Exception:
        pass

    inicio_sesion = datetime.now()
    try:
        iniciar_monitoreo_en_oracle(inicio_sesion)
    except Exception:
        pass

    global error_activo, error_inicio, ultimo_error_msg, last_mon_update
    last_mon_update = inicio_sesion

    while True:
        try:
            async with Client(url=URL) as client:
                logger.info("Conectado al servidor OPC UA")

                if error_activo:
                    try:
                        registrar_error_conexion(error_inicio, datetime.now(), ultimo_error_msg)
                    except Exception:
                        pass
                    error_activo = False
                    error_inicio = None
                    ultimo_error_msg = ""

                ncache = NodeCache(client).build_for_groups(NODE_GROUPS)

                while True:
                    await asyncio.gather(*[
                        leer_grupo_si_falla(client, grupo, ncache.for_group(grupo))
                        for grupo in NODE_GROUPS
                    ])
                    await asyncio.sleep(INTERVALO_SEGUNDOS)

                    ahora = datetime.now()
                    if (ahora - last_mon_update).total_seconds() >= 60:
                        try:
                            actualizar_fin_monitoreo(ahora)
                        except Exception:
                            pass
                        last_mon_update = ahora

        except Exception as e:
            if not error_activo:
                error_activo = True
                error_inicio = datetime.now()
                ultimo_error_msg = str(e)
                logger.error(f"Conexión caída: {ultimo_error_msg}")
            await asyncio.sleep(2)  # retry

# ================== HILO DE MONITOREO ==================
def thread_monitor():
    """Función de entrada para el hilo de monitoreo. Inicia el bucle asíncrono."""
    asyncio.run(main())

# ================== ESTILOS ==================
# Definición de los estilos CSS para el dashboard de Dash.
BG = "#0f172a"; CARD = "#111827"; TEXT = "#e5e7eb"; MUTED = "#9ca3af"
ACCENT = "#22d3ee"; ACCENT_SOFT = "rgba(34,211,238,0.15)"

container_style = {
    "backgroundColor": BG, "minHeight": "100vh", "padding": "1px 22px",
    "fontFamily": "Inter, system-ui, -apple-system, Segoe UI, Roboto, Ubuntu, Cantarell, Noto Sans, Helvetica, Arial"
}

title_style = {"color": TEXT, "fontSize": "24px", "marginBottom": "10px", "fontWeight": 700}
card_style = {"backgroundColor": CARD, "borderRadius": "14px", "padding": "14px 16px",
              "boxShadow": "0 4px 16px rgba(0,0,0,0.35)", "border": f"1px solid {ACCENT_SOFT}"}

table_card_style = {**card_style, "padding": "6px 8px"}

# ================== DASHBOARD (Dash) ==================
# Definición del layout del dashboard.
# Incluye dos tablas (fallas activas y errores) y un componente dcc.Interval
# para actualizar los datos automáticamente.
app = Dash(__name__)
app.layout = html.Div([
    dcc.Store(id="store-snapshot", storage_type="local"),

    html.Div([
        # ===== Columna IZQUIERDA (75%): Fallas Activas =====
        html.Div([
            html.H3("🛠️ Fallas Activas", style=title_style),
            html.Div([
                dash_table.DataTable(
                    id="tabla-fallas",
                    columns=[
                        {"name": "Equipo", "id": "Equipo"},
                        {"name": "GMF", "id": "GMF"},
                        {"name": "Descripción", "id": "Descripcion"},
                        {"name": "Tiempo (mm:ss)", "id": "TiempoFmt"},
                        {"name": "Tiempo OFF (seg)", "id": "TiempoOffSeg"},
                    ],
                    data=[],
                    sort_action="native",
                    page_action="native",
                    page_size=12,
                    style_as_list_view=True,
                    style_table={
                        "height": "78vh",
                        "overflowY": "auto",
                        "border": f"1px solid {ACCENT_SOFT}",
                        "borderRadius": "12px"
                    },
                    style_header={
                        "backgroundColor": "#111827",
                        "color": TEXT,
                        "fontWeight": "700",
                        "borderBottom": f"2px solid {ACCENT_SOFT}",
                        "position": "sticky", "top": 0, "zIndex": 1
                    },
                    style_cell={
                        "backgroundColor": "#0b1220",
                        "color": TEXT,
                        "padding": "8px 10px",
                        "border": "0px",
                        "fontSize": "22px",
                        "whiteSpace": "nowrap",
                        "textOverflow": "ellipsis",
                        "maxWidth": 380
                    },
                    style_data_conditional=[
                        {"if": {"filter_query": "{TiempoOffSeg} >= 60"},
                         "backgroundColor": "#783c3c", "color": "#ffdddd", "fontWeight": "700",
                         "borderLeft": "4px solid #ff4d4f"},
                        {"if": {"filter_query": "{TiempoOffSeg} >= 30 && {TiempoOffSeg} < 60"},
                         "backgroundColor": "#3f3b1d", "color": "#fff1b8", "fontWeight": "700",
                         "borderLeft": "4px solid #facc15"},
                        {"if": {"row_index": "odd", "filter_query": "{TiempoOffSeg} < 30"},
                         "backgroundColor": "#0d1424"},
                        {"if": {"state": "active"}, "border": f"1px solid {ACCENT_SOFT}"},
                        {"if": {"column_id": "TiempoFmt"}, "fontWeight": "700", "color": ACCENT},
                    ],
                ),
                html.Div(id="last-update", style={"color": MUTED, "fontSize": "12px", "marginTop": "8px"}),
            ], style=table_card_style),
        ], style={"flexBasis": "75%", "minWidth": 0}),

        # ===== Columna DERECHA (25%): Errores de conexión =====
        html.Div([
            html.H3("🚨 Errores de conexión", style=title_style),
            html.Div([
                dash_table.DataTable(
                    id="tabla-errores",
                    columns=[
                        {"name": "Grupo", "id": "Grupo"},
                        {"name": "Desde", "id": "Desde"},
                        {"name": "Hace", "id": "Hace"},
                        {"name": "Msj", "id": "Mensaje"},
                    ],
                    data=[],
                    sort_action="native",
                    page_action="native",
                    page_size=8,
                    style_as_list_view=True,
                    style_table={
                        "height": "78vh",
                        "overflowY": "auto",
                        "border": f"1px solid {ACCENT_SOFT}",
                        "borderRadius": "12px"
                    },
                    style_header={
                        "backgroundColor": "#111827",
                        "color": TEXT,
                        "fontWeight": "700",
                        "borderBottom": f"2px solid {ACCENT_SOFT}",
                        "position": "sticky", "top": 0, "zIndex": 1
                    },
                    style_cell={
                        "backgroundColor": "#0b1220",
                        "color": TEXT,
                        "padding": "8px 10px",
                        "border": "0px",
                        "fontSize": "18px",
                        "whiteSpace": "nowrap",
                        "textOverflow": "ellipsis",
                        "maxWidth": 600
                    },
                    style_data_conditional=[
                        {"if": {"row_index": "odd"}, "backgroundColor": "#0d1424"},
                        {"if": {"state": "active"}, "border": f"1px solid {ACCENT_SOFT}"},
                        {"if": {"column_id": "Hace"}, "fontWeight": "700", "color": ACCENT},
                        {"if": {"filter_query": "{Mensaje} contains 'Error' || {Mensaje} contains 'Failed'"},
                         "backgroundColor": "#5b1a1a", "color": "#ffe5e5",
                         "borderLeft": "4px solid #ef4444", "fontWeight": "700"},
                    ],
                ),
            ], style=table_card_style),
        ], style={"flexBasis": "25%", "minWidth": 0}),

    ], style={
        "display": "flex",
        "flexDirection": "row",
        "gap": "16px",
        "alignItems": "stretch",
        "width": "100%",
    }),

    dcc.Interval(id="intervalo", interval=500, n_intervals=0),
], style=container_style)


# ================== HELPERS & CALLBACKS ==================
def _fmt_mmss(segundos: int) -> str:
    """Convierte segundos en formato 'minutos:segundos'."""
    m, s = divmod(segundos, 60)
    return f"{m:02d}:{s:02d}"

@app.callback(
    Output("tabla-fallas", "data"),
    Output("tabla-errores", "data"),
    Output("last-update", "children"),
    Input("intervalo", "n_intervals")
)
def update_dashboard(n):
    """
    Función de callback de Dash que se ejecuta en cada intervalo.
    Actualiza los datos de las tablas del dashboard con el estado actual
    de las fallas y los errores de conexión.
    """
    with lock:
        # Crea la lista de fallas activas para la tabla principal.
        fallas_data = []
        for gmf_id, inicio_falla in tiempos_gmf.items():
            grupo, codigo = gmf_id.split(".")
            ahora = datetime.now()
            segundos_activos = int((ahora - inicio_falla).total_seconds())
            
            # Si el equipo está en 'continuous=0', usa el tiempo congelado.
            if not last_continuous_state.get(grupo):
                segundos_activos = tiempos_congelados.get(gmf_id, segundos_activos)

            fallas_data.append({
                "Equipo": grupo,
                "GMF": codigo,
                "Descripcion": _desc_gmf(grupo, codigo),
                "TiempoFmt": _fmt_mmss(segundos_activos),
                "TiempoOffSeg": segundos_activos
            })
        
        # Crea la lista de errores para la tabla de errores.
        errores_data = []
        for grupo, esta_activo in error_grupo_activo.items():
            if esta_activo:
                inicio_error = error_grupo_inicio[grupo]
                segundos_error = int((datetime.now() - inicio_error).total_seconds())
                errores_data.append({
                    "Grupo": grupo,
                    "Desde": inicio_error.strftime("%H:%M:%S"),
                    "Hace": _fmt_mmss(segundos_error),
                    "Mensaje": error_grupo_msg[grupo]
                })

        # Incluye el error de conexión global si está activo.
        if error_activo:
            segundos_error = int((datetime.now() - error_inicio).total_seconds())
            errores_data.insert(0, {
                "Grupo": "Conexión Global",
                "Desde": error_inicio.strftime("%H:%M:%S"),
                "Hace": _fmt_mmss(segundos_error),
                "Mensaje": ultimo_error_msg
            })

    # Actualiza la etiqueta de la última actualización.
    last_update_text = f"Última actualización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    return fallas_data, errores_data, last_update_text

# ================== EJECUCIÓN ==================
if __name__ == '__main__':
    # Esta es la entrada principal del script.
    # Inicia el bucle de monitoreo en un hilo separado del servidor de Dash,
    # y luego inicia el servidor web.
    mon_thread = threading.Thread(target=thread_monitor, daemon=True)
    mon_thread.start()

    app.run_server(debug=False)