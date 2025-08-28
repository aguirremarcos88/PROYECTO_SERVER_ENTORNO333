from __future__ import annotations

import asyncio
import logging
import os
import sys
import threading
import uuid
from contextlib import contextmanager, suppress
from dataclasses import dataclass, field
from datetime import datetime, time, timedelta
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Optional

import requests
from asyncua import Client, ua
from dash import Dash, dcc, html
from dash import dash_table
from dash.dependencies import Input, Output
from flask import Response
from openpyxl import load_workbook

# ======================================================================================
# Importación y configuración inicial
# ======================================================================================
# Se importan las librerías necesarias. Se incluye un manejo de errores para los
# drivers de Oracle y las excepciones de `asyncua` para asegurar compatibilidad.
# La carga de las variables de entorno desde un archivo `.env` se maneja
# de forma opcional para la configuración.

# ===== DB driver preferido =====
try:
    import cx_Oracle as db  # type: ignore
except ModuleNotFoundError:  # pragma: no cover
    import oracledb as db  # type: ignore

# ===== Excepción opcua segun versión =====
try:
    from asyncua.ua.uaerrors import BadNodeIdUnknown  # type: ignore
except Exception:  # pragma: no cover
    try:
        from asyncua.ua.uaerrors._auto import BadNodeIdUnknown  # type: ignore
    except Exception:  # pragma: no cover
        class BadNodeIdUnknown(Exception):
            pass

# ===== (Opcional) .env =====
try:  # pragma: no cover
    from dotenv import load_dotenv  # type: ignore
except Exception:  # pragma: no cover
    load_dotenv = None  # type: ignore

BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
if load_dotenv:
    load_dotenv(BASE_DIR / ".env")

# --------------------------------------------------------------------------------------
# Configuración del entorno
# --------------------------------------------------------------------------------------
# Se definen funciones de ayuda para leer variables de entorno con valores por
# defecto, lo que hace que la configuración sea más robusta y fácil de usar.

def env_str(name: str, default: str) -> str:
    """Obtiene una variable de entorno como string, con un valor por defecto."""
    return os.getenv(name, default)


def env_bool(name: str, default: bool) -> bool:
    """Obtiene una variable de entorno como booleano, con un valor por defecto.
    Acepta '1', 'true', 't', 'yes', 'y' como `True`.
    """
    v = os.getenv(name)
    if v is None:
        return default
    return v.strip().lower() in {"1", "true", "t", "yes", "y"}


def env_int(name: str, default: int) -> int:
    """Obtiene una variable de entorno como entero, con un valor por defecto y manejo de errores."""
    try:
        return int(os.getenv(name, str(default)))
    except Exception:
        return default


# --------------------------------------------------------------------------------------
# Clase de configuración centralizada
# --------------------------------------------------------------------------------------
# La clase `Config` agrupa todas las configuraciones del sistema, desde la conexión OPC UA
# hasta las opciones de Telegram y el servidor web. El uso de `dataclass` y
# `field(default_factory)` mejora la legibilidad y gestión de la configuración.
# La configuración se carga una sola vez al inicio en la variable global `CFG`.

@dataclass(frozen=True)
class Config:
    url: str = env_str("OPCUA_URL", "opc.tcp://localhost:52250/freeopcua/server/")
    node_groups: list[str] = field(
        default_factory=lambda: [g.strip() for g in env_str("NODE_GROUPS", "BOMBAS_SH,CONVEYOR_SH").split(",") if g.strip()]
    )
    ns: int = env_int("OPCUA_NS", 2)
    start_hex: int = int(env_str("START_HEX", "0x80"), 16)
    end_hex: int = int(env_str("END_HEX", "0xFF"), 16)

    sampling_ms: int = env_int("SAMPLING_MS", 100)
    sub_queue_size: int = env_int("SUB_QUEUE_SIZE", 1000)
    sub_keepalive_count: int = env_int("SUB_KEEPALIVE_COUNT", 100)
    sub_lifetime_count: int = env_int("SUB_LIFETIME_COUNT", env_int("SUB_KEEPALIVE_COUNT", 100) * 10)

    ui_interval_ms: int = int(float(env_str("INTERVALO_SEGUNDOS", "0.5")) * 1000)

    log_level: str = env_str("LOG_LEVEL", "INFO").upper()
    log_file: str = str((BASE_DIR / "monitor.log").resolve())

    telegram_enabled: bool = env_bool("TELEGRAM_ENABLED", True)
    telegram_token: Optional[str] = os.getenv("TELEGRAM_TOKEN")
    telegram_chat_ids: list[str] = field(
        default_factory=lambda: [x.strip() for x in os.getenv("TELEGRAM_CHAT_IDS", "1600754452,5015132163").split(",") if x.strip()]
    )
    telegram_threshold_seconds: int = env_int("TELEGRAM_THRESHOLD_SECONDS", 60)
    telegram_restablecido_unico: bool = env_bool("TELEGRAM_RESTABLECIDO_UNICO", True)

    host: str = env_str("HOST", "0.0.0.0")
    port: int = env_int("PORT", 8050)


CFG = Config()

# --------------------------------------------------------------------------------------
# Configuración del Logging
# --------------------------------------------------------------------------------------
# Se configura un logger global para toda la aplicación. Utiliza `RotatingFileHandler`
# para gestionar el tamaño de los archivos de log y un `StreamHandler` para mostrar
# la salida en la consola. Esto permite tener un registro de los eventos y errores.

logger = logging.getLogger("opcua-monitor")
logger.setLevel(CFG.log_level)
logger.propagate = False
fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s", "%Y-%m-%d %H:%M:%S")
logger.handlers.clear()
fh = RotatingFileHandler(CFG.log_file, maxBytes=1_000_000, backupCount=3, encoding="utf-8")
fh.setFormatter(fmt)
sh = logging.StreamHandler(sys.stdout)
sh.setFormatter(fmt)
logger.addHandler(fh)
logger.addHandler(sh)

# --------------------------------------------------------------------------------------
# Cliente Oracle
# --------------------------------------------------------------------------------------
# Se utiliza un `context manager` (`@contextmanager`) para simplificar la conexión
# y el manejo de cursores de Oracle. Esto garantiza que las conexiones se cierren
# correctamente incluso si ocurren errores, evitando la repetición de código.
# La clase `OracleRepo` centraliza las operaciones de base de datos, como la
# creación de tablas y la inserción de registros.

def _oracle_conn():
    """Establece una conexión a la base de datos de Oracle."""
    user = os.getenv("ORACLE_USER")
    pwd = os.getenv("ORACLE_PWD")
    dsn = os.getenv("ORACLE_DSN")
    if not all([user, pwd, dsn]):
        raise RuntimeError("Faltan ORACLE_USER/ORACLE_PWD/ORACLE_DSN en .env")
    return db.connect(user=user, password=pwd, dsn=dsn)


@contextmanager
def oracle_cursor():
    """Context manager para manejar el ciclo de vida de una conexión y un cursor de Oracle."""
    conn = _oracle_conn()
    cur = conn.cursor()
    try:
        yield cur
        conn.commit()
    finally:
        with suppress(Exception):
            cur.close()
            conn.close()


class OracleRepo:
    """Clase para interactuar con la base de datos de Oracle, encapsulando las operaciones."""
    def ensure_schema(self):
        """Crea las tablas `MONITOREO_OPCUA` y añade la columna `EVENT_ID` si no existen."""
        try:
            with oracle_cursor() as cur:
                # Código SQL para crear la tabla MONITOREO_OPCUA (si no existe)
                cur.execute(
                    """
                    BEGIN
                        EXECUTE IMMEDIATE '
                            CREATE TABLE MONITOREO_OPCUA (
                                SESSION_ID      VARCHAR2(64),
                                TIPO            VARCHAR2(30),
                                GRUPO           VARCHAR2(50),
                                INICIO          TIMESTAMP,
                                FIN             TIMESTAMP,
                                MENSAJE         VARCHAR2(4000)
                            )';
                    EXCEPTION WHEN OTHERS THEN
                        IF SQLCODE != -955 THEN RAISE; END IF;
                    END;"""
                )
                # Intentar agregar EVENT_ID si no existe (ORA-01430)
                with suppress(Exception):
                    cur.execute(
                        """
                        BEGIN
                            EXECUTE IMMEDIATE 'ALTER TABLE MONITOREO_OPCUA ADD (EVENT_ID VARCHAR2(64))';
                        EXCEPTION WHEN OTHERS THEN
                            IF SQLCODE != -1430 THEN RAISE; END IF;
                        END;"""
                    )
        except Exception as e:  # pragma: no cover
            logger.error(f"Error asegurando tablas Oracle: {e}")

    def start_monitoring(self, session_id: str, inicio: datetime):
        """Inserta un registro para marcar el inicio de una sesión de monitoreo."""
        with suppress(Exception):
            with oracle_cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
                    VALUES (:sid, 'MONITOREO', NULL, :ini, :fin, :msg)
                    """,
                    dict(sid=session_id, ini=inicio, fin=inicio, msg=None),
                )

    def tick_monitoring(self, session_id: str, fin: datetime):
        """Actualiza la marca de tiempo de la sesión de monitoreo para indicar que sigue activa."""
        with suppress(Exception):
            with oracle_cursor() as cur:
                cur.execute(
                    """UPDATE MONITOREO_OPCUA SET FIN = :fin WHERE SESSION_ID = :sid AND TIPO = 'MONITOREO'""",
                    dict(fin=fin, sid=session_id),
                )

    def insert_error_conn(self, session_id: str, inicio: datetime, fin: datetime, msg: str):
        """Inserta un registro de error de conexión en la base de datos."""
        with suppress(Exception):
            with oracle_cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
                    VALUES (:sid, 'ERROR_CONEXION', NULL, :ini, :fin, :msg)
                    """,
                    dict(sid=session_id, ini=inicio, fin=fin, msg=(msg or "")[:3999]),
                )

    def open_group_error(self, session_id: str, grupo: str, inicio: datetime, msg: str, eid: Optional[str]):
        """Crea un registro para una falla de grupo, con `FIN` en `NULL`."""
        with suppress(Exception):
            with oracle_cursor() as cur:
                try:
                    cur.execute(
                        """
                        INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE, EVENT_ID)
                        VALUES (:sid, 'ERROR_GRUPO', :grp, :ini, NULL, :msg, :eid)
                        """,
                        dict(sid=session_id, grp=grupo, ini=inicio, msg=(msg or "")[:3999], eid=eid),
                    )
                except Exception:
                    cur.execute(
                        """
                        INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
                        VALUES (:sid, 'ERROR_GRUPO', :grp, :ini, NULL, :msg)
                        """,
                        dict(sid=session_id, grp=grupo, ini=inicio, msg=(msg or "")[:3999]),
                    )

    def close_group_error(self, session_id: str, grupo: str, fin: datetime, eid: Optional[str]):
        """Actualiza un registro de falla de grupo con la fecha y hora de finalización."""
        with suppress(Exception):
            with oracle_cursor() as cur:
                if eid:
                    cur.execute(
                        """
                        UPDATE MONITOREO_OPCUA SET FIN = :fin
                          WHERE SESSION_ID = :sid AND TIPO = 'ERROR_GRUPO' AND GRUPO = :grp AND EVENT_ID = :eid AND FIN IS NULL
                        """,
                        dict(fin=fin, sid=session_id, grp=grupo, eid=eid),
                    )
                else:
                    cur.execute(
                        """
                        UPDATE MONITOREO_OPCUA SET FIN = :fin
                          WHERE SESSION_ID = :sid AND TIPO = 'ERROR_GRUPO' AND GRUPO = :grp AND FIN IS NULL
                        """,
                        dict(fin=fin, sid=session_id, grp=grupo),
                    )


ORACLE = OracleRepo()
ORACLE.ensure_schema()

# --------------------------------------------------------------------------------------
# Cliente Telegram
# --------------------------------------------------------------------------------------
# La clase `TelegramClient` encapsula la lógica para enviar mensajes a través de la API
# de Telegram. Se inicializa con los parámetros de configuración y tiene un método
# `send` que maneja el envío, incluyendo el manejo de errores.

class TelegramClient:
    """Clase para enviar mensajes a través de la API de Telegram."""
    def __init__(self, enabled: bool, token: Optional[str], chat_ids: list[str]) -> None:
        self.enabled = enabled and bool(token) and any(cid.strip().lstrip("-").isdigit() for cid in chat_ids)
        self.token = token
        self.chat_ids = [c.strip() for c in chat_ids if c.strip()]
        self.session = requests.Session()

    def send(self, text: str):
        """Envía un mensaje a todos los `chat_ids` configurados."""
        if not self.enabled or not self.token:
            return
        url = f"https://api.telegram.org/bot{self.token}/sendMessage"
        for cid in self.chat_ids:
            payload = dict(chat_id=cid, text=text, parse_mode="HTML", disable_web_page_preview=True)
            try:
                r = self.session.post(url, json=payload, timeout=(3, 8))
                if r.status_code != 200:
                    logger.error(f"Telegram error chat_id={cid}: {r.status_code} {r.text}")
            except Exception as e:  # pragma: no cover
                logger.error(f"Telegram excepción chat_id={cid}: {e}")


TELEGRAM = TelegramClient(CFG.telegram_enabled, CFG.telegram_token, CFG.telegram_chat_ids)

# --------------------------------------------------------------------------------------
# Gestión del estado de la aplicación
# --------------------------------------------------------------------------------------
# La clase `State` es un `dataclass` que centraliza todas las variables de estado
# de la aplicación. Esto incluye los tiempos de las fallas, la información de Telegram,
# y los estados de error. El uso de un `threading.Lock` asegura que las
# operaciones sobre el estado sean seguras en entornos de múltiples hilos.

@dataclass
class State:
    """Clase para gestionar el estado de la aplicación de forma centralizada y segura."""
    node_groups: list[str]
    session_id: str = field(default_factory=lambda: uuid.uuid4().hex)

    lock: threading.Lock = field(default_factory=threading.Lock)

    tiempos_gmf: dict[str, datetime] = field(default_factory=dict)
    tiempos_congelados: dict[str, int] = field(default_factory=dict)
    acumulado_continuous: dict[str, timedelta] = field(default_factory=dict)
    ultima_marca_continuous: dict[str, datetime] = field(default_factory=dict)
    descripciones_gmf: dict[str, dict[str, str]] = field(default_factory=dict)

    tg_enviado: dict[str, bool] = field(init=False)
    tg_restaurado_enviado: dict[str, bool] = field(init=False)

    last_mon_update: datetime = field(default_factory=lambda: datetime.min)

    error_activo: bool = False
    error_inicio: Optional[datetime] = None
    ultimo_error_msg: str = ""
    error_grupo_activo: dict[str, bool] = field(init=False)
    error_grupo_inicio: dict[str, Optional[datetime]] = field(init=False)
    error_grupo_msg: dict[str, str] = field(init=False)
    error_grupo_event_id: dict[str, Optional[str]] = field(init=False)

    current_continuous: dict[str, bool] = field(init=False)
    state_cont_known: dict[str, bool] = field(init=False)
    last_word_values: dict[tuple[str, str], int] = field(default_factory=dict)

    def __post_init__(self):
        """Inicializa los diccionarios de estado después de la construcción del objeto."""
        self.tg_enviado = {g: False for g in self.node_groups}
        self.tg_restaurado_enviado = {g: False for g in self.node_groups}
        self.error_grupo_activo = {g: False for g in self.node_groups}
        self.error_grupo_inicio = {g: None for g in self.node_groups}
        self.error_grupo_msg = {g: "" for g in self.node_groups}
        self.error_grupo_event_id = {g: None for g in self.node_groups}
        self.current_continuous = {g: True for g in self.node_groups}
        self.state_cont_known = {g: False for g in self.node_groups}

    # ---- helpers ----
    def marcar_error_grupo(self, grupo: str, mensaje: str, inicio: Optional[datetime] = None):
        """Registra un error para un grupo específico en el estado y en la base de datos."""
        ahora = inicio or datetime.now()
        with self.lock:
            ya_activo = self.error_grupo_activo.get(grupo, False)
            self.error_grupo_activo[grupo] = True
            if not ya_activo:
                self.error_grupo_inicio[grupo] = ahora
                eid = uuid.uuid4().hex
                self.error_grupo_event_id[grupo] = eid
            else:
                eid = self.error_grupo_event_id.get(grupo)
            self.error_grupo_msg[grupo] = mensaje
        if not ya_activo:
            ORACLE.open_group_error(self.session_id, grupo, ahora, mensaje, eid)
            logger.warning(f"ERROR_GRUPO abierto: {grupo} @ {ahora} | {mensaje} | event_id={eid}")

    def limpiar_error_grupo(self, grupo: str, fin: Optional[datetime] = None):
        """Cierra un error de grupo activo en el estado y en la base de datos."""
        ahora = fin or datetime.now()
        with self.lock:
            estaba = self.error_grupo_activo.get(grupo, False)
            eid = self.error_grupo_event_id.get(grupo)
            self.error_grupo_activo[grupo] = False
            self.error_grupo_event_id[grupo] = None
        if estaba:
            ORACLE.close_group_error(self.session_id, grupo, ahora, eid)
            logger.info(f"ERROR_GRUPO cerrado: {grupo} @ {ahora} | event_id={eid}")


STATE = State(CFG.node_groups)

# --------------------------------------------------------------------------------------
# Utilidades
# --------------------------------------------------------------------------------------
# Funciones de ayuda para tareas comunes, como la construcción de Node IDs,
# la determinación del turno de trabajo, el formateo de tiempo y la
# normalización de códigos.

def nid(s: str) -> str:
    """Genera un NodeId de OPC UA a partir de un string de nombre de nodo."""
    return f"ns={CFG.ns};s={s}"


def determinar_turno(fecha: datetime) -> str:
    """Determina el turno de trabajo ('MAÑANA', 'TARDE', 'NOCHE') a partir de la hora."""
    hora = fecha.time()
    if time(6, 5) <= hora <= time(14, 30):
        return "MAÑANA"
    if time(14, 39) <= hora <= time(22, 54):
        return "TARDE"
    if time(23, 13) <= hora <= time(23, 59) or time(0, 0) <= hora <= time(5, 55):
        return "NOCHE"
    return "FUERA DE TURNO"


def _fmt_mmss(segundos: int) -> str:
    """Formatea segundos en formato `MM:SS`."""
    m, s = divmod(max(0, int(segundos)), 60)
    return f"{m:02d}:{s:02d}"


def _normalize_code(codigo) -> str:
    """Normaliza un código de falla (e.g., de Excel) a un formato consistente."""
    if codigo is None:
        return ""
    if isinstance(codigo, (int, float)):
        s = f"{codigo}"
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s.strip().upper()
    return str(codigo).strip().upper()


def _desc_gmf(grupo: str, gmf: str) -> str:
    """Busca la descripción de un código GMF en el estado de la aplicación."""
    return STATE.descripciones_gmf.get(grupo, {}).get(gmf.strip().upper(), "Sin descripción")


def _msg_telegram(grupo: str, segundos: int, activos: list[str]) -> str:
    """Crea el mensaje formateado para enviar a Telegram."""
    import html as html_escape

    if not activos:
        return f"⚠️ {html_escape.escape(grupo)} en falla, hace {segundos} s."
    if len(activos) == 1:
        g = activos[0]
        d = _desc_gmf(grupo, g)
        return f"⚠️ {html_escape.escape(grupo)} con falla {html_escape.escape(g)}: {html_escape.escape(d)} , hace {segundos} s."
    partes = [f"{html_escape.escape(g)}: {html_escape.escape(_desc_gmf(grupo, g))}" for g in activos]
    listado = " // ".join(partes[:10])  # antes: "; ".join(partes[:10])
    extra = "" if len(partes) <= 10 else f" (+{len(partes)-10} más)"
    return f"⚠️ {html_escape.escape(grupo)} con fallas <b>{listado}{extra}</b> , hace {segundos} s."


# --------------------------------------------------------------------------------------
# Carga de descripciones (Excel)
# --------------------------------------------------------------------------------------
# La función `cargar_descripciones` lee archivos de Excel que contienen los códigos
# de fallas (GMFs) y sus descripciones. Esta información se guarda en el objeto `STATE`
# para su uso posterior, por ejemplo, en los mensajes de Telegram.

ARCHIVOS_DESCRIPCIONES = {
    "BOMBAS_SH": str((BASE_DIR / "GMFs_de_Equipos" / "GMF_BOMBAS_SH.xlsx").resolve()),
    "CONVEYOR_SH": str((BASE_DIR / "GMFs_de_Equipos" / "GMF_CONVEYOR_SH.xlsx").resolve()),
}


def cargar_descripciones():
    """Carga las descripciones de los códigos de falla desde archivos de Excel."""
    for grupo, ruta in ARCHIVOS_DESCRIPCIONES.items():
        STATE.descripciones_gmf[grupo] = {}
        try:
            wb = load_workbook(ruta, data_only=True)
            hoja = wb.active
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                codigo_raw, descripcion = fila[:2]
                if codigo_raw and descripcion:
                    codigo = _normalize_code(codigo_raw)
                    if not codigo.startswith("GMF"):
                        codigo = "GMF" + codigo
                    STATE.descripciones_gmf[grupo][codigo] = str(descripcion).strip()
            logger.info(f"[Desc] {grupo}: {len(STATE.descripciones_gmf[grupo])} códigos desde {ruta}")
        except Exception as e:  # pragma: no cover
            logger.error(f"Error cargando descripciones {grupo}: {e}")


# --------------------------------------------------------------------------------------
# Exportación a ORACLE
# --------------------------------------------------------------------------------------
# La función `exportar_a_oracle` se encarga de persistir los registros de fallas
# en la base de datos, creando tablas específicas para cada grupo de nodos (`HISTORIAL_FALLAS_*`).
# Esto permite mantener un historial completo de las fallas.

def exportar_a_oracle(grupo: str, gmf: str, descripcion: str, inicio: datetime, fin: datetime, acumulado: int):
    """Exporta los datos de una falla específica a la base de datos de Oracle."""
    try:
        tabla = f"HISTORIAL_FALLAS_{grupo.upper()}"
        with oracle_cursor() as cur:
            # Código SQL para crear la tabla de historial (si no existe)
            cur.execute(
                f"""
                BEGIN
                    EXECUTE IMMEDIATE '
                        CREATE TABLE {tabla} (
                            EQUIPO             VARCHAR2(50),
                            FALLA              VARCHAR2(50),
                            DESCRIPCION        VARCHAR2(255),
                            INICIO_FALLA       TIMESTAMP,
                            FIN_FALLA          TIMESTAMP,
                            TIEMPO_OFF_SEG     NUMBER,
                            TURNO              VARCHAR2(20)
                        )';
                EXCEPTION WHEN OTHERS THEN
                    IF SQLCODE != -955 THEN RAISE; END IF;
                END;"""
            )
            cur.execute(
                f"""
                INSERT INTO {tabla} (EQUIPO, FALLA, DESCRIPCION, INICIO_FALLA, FIN_FALLA, TIEMPO_OFF_SEG, TURNO)
                VALUES (:1, :2, :3, :4, :5, :6, :7)
                """,
                (grupo, gmf, descripcion, inicio, fin, acumulado, determinar_turno(inicio)),
            )
        logger.info(f"Exportado a Oracle: {grupo} | {gmf} | {acumulado} seg")
    except Exception as e:  # pragma: no cover
        logger.error(f"Error exportando a Oracle: {e}")


# --------------------------------------------------------------------------------------
# Exportación de fallas "congeladas"
# --------------------------------------------------------------------------------------
# La función `_exportar_congeladas_de_grupo` maneja un caso específico: cuando el valor
# de `CONTINUOUS` pasa de `0` a `1`, se exportan las fallas que estaban activas
# mientras el monitoreo estaba "congelado".

def _exportar_congeladas_de_grupo(grupo: str, ahora: datetime):
    """Exporta a Oracle las fallas "congeladas" de un grupo (cuando CONTINUOUS estuvo en 0)
    y resetea el acumulado OFF del grupo.
    """
    with STATE.lock:
        pendientes = [
            (k.split(".")[1], seg)
            for k, seg in STATE.tiempos_congelados.items()
            if k.startswith(grupo + ".")
        ]
        segundos_off = int(STATE.acumulado_continuous.get(grupo, timedelta()).total_seconds())

    for gmf, seg in pendientes:
        try:
            inicio = ahora - timedelta(seconds=int(seg))
            descripcion = _desc_gmf(grupo, gmf)
            exportar_a_oracle(grupo, gmf, descripcion, inicio, ahora, segundos_off)
            with STATE.lock:
                STATE.tiempos_congelados.pop(f"{grupo}.{gmf}", None)
        except Exception as e:  # pragma: no cover
            logger.error(f"Error exportando congelada {grupo}.{gmf}: {e}")

    # Reset OFF tras exportar todas
    with STATE.lock:
        STATE.acumulado_continuous[grupo] = timedelta(0)

# Alias por compatibilidad
exportar_congeladas_de_grupo = _exportar_congeladas_de_grupo

# --------------------------------------------------------------------------------------
# Node cache
# --------------------------------------------------------------------------------------
# La clase `NodeCache` precarga los nodos OPC UA de interés en una caché para
# evitar búsquedas repetitivas. El método `_validate_gmfs` verifica que los nodos
# GMF definidos en el rango hexadecimal realmente existan en el servidor OPC UA.

class NodeCache:
    """Clase para almacenar en caché los objetos de nodo de OPC UA para un acceso eficiente."""
    def __init__(self, client: Client):
        self.client = client
        self.cache: dict[str, dict] = {}

    def build_for_groups(self, grupos: list[str]):
        """Construye la caché de nodos para los grupos definidos."""
        for grupo in grupos:
            total_fault = self.client.get_node(nid(f"{grupo}.TOTAL_FAULT"))
            continuous = self.client.get_node(nid(f"{grupo}.CONTINUOUS"))
            total_start = self.client.get_node(nid(f"{grupo}.TOTAL_START"))
            hex_suffixes = [f"{i:X}" for i in range(CFG.start_hex, CFG.end_hex + 1)]
            node_ids = [ua.NodeId.from_string(nid(f"{grupo}.GMF{suf}")) for suf in hex_suffixes]
            gmf_nodes = [self.client.get_node(n) for n in node_ids]
            self.cache[grupo] = {
                "total_fault": total_fault,
                "continuous": continuous,
                "total_start": total_start,
                "gmf_nodes": gmf_nodes,
                "gmf_suffixes": hex_suffixes,
                "gmf_validated": False,
            }
        return self

    def for_group(self, grupo: str) -> dict:
        """Devuelve la caché de nodos para un grupo específico."""
        return self.cache[grupo]


async def _validate_gmfs(client: Client, grupo: str, ncache: dict):
    """Valida la existencia de los nodos GMF en el servidor OPC UA."""
    valid_nodes, valid_suffixes = [], []
    for node, suf in zip(ncache["gmf_nodes"], ncache["gmf_suffixes"]):
        try:
            await node.read_value()
            valid_nodes.append(node)
            valid_suffixes.append(suf)
        except BadNodeIdUnknown:
            logger.debug(f"{grupo}: GMF{suf} no existe, se omite.")
        except Exception:
            logger.debug(f"{grupo}: GMF{suf} no legible por ahora, se omite.")
    ncache["gmf_nodes"], ncache["gmf_suffixes"], ncache["gmf_validated"] = valid_nodes, valid_suffixes, True


# --------------------------------------------------------------------------------------
# Subscription handler (Manejo de suscripciones)
# --------------------------------------------------------------------------------------
# La clase `SubHandler` implementa la lógica para procesar los cambios de datos recibidos
# del servidor OPC UA. Contiene métodos para manejar el estado de `CONTINUOUS` y
# los valores de los códigos GMF, abriendo y cerrando fallas según corresponda.

class SubHandler:
    """Manejador de notificaciones de cambio de datos de las suscripciones OPC UA."""
    def __init__(self, client: Client, ncache: NodeCache):
        self.client = client
        self.ncache = ncache

    async def _cerrar_falla(self, grupo: str, gmf_codigo: str, ahora: datetime):
        """Cierra una falla de GMF y exporta los datos a la base de datos."""
        try:
            descripcion = _desc_gmf(grupo, gmf_codigo)
            with STATE.lock:
                inicio_falla = STATE.tiempos_gmf.get(f"{grupo}.{gmf_codigo}")
                acumulado = int(STATE.acumulado_continuous.get(grupo, timedelta()).total_seconds())
                cont_ok = bool(STATE.current_continuous.get(grupo, True)) and STATE.state_cont_known.get(grupo, False)

            if cont_ok:
                exportar_a_oracle(grupo, gmf_codigo, descripcion, inicio_falla, ahora, acumulado)
                with STATE.lock:
                    STATE.acumulado_continuous[grupo] = timedelta(0)

                if (acumulado > CFG.telegram_threshold_seconds) and (not STATE.tg_restaurado_enviado.get(grupo, False)):
                    TELEGRAM.send(f'✅ Equipo "{grupo}" restablecido. Tiempo acumulado {int(max(0, acumulado))} s')
                    if CFG.telegram_restablecido_unico:
                        STATE.tg_restaurado_enviado[grupo] = True

                with STATE.lock:
                    STATE.tiempos_gmf.pop(f"{grupo}.{gmf_codigo}", None)
                    STATE.tiempos_congelados.pop(f"{grupo}.{gmf_codigo}", None)
            else:
                with STATE.lock:
                    clave = f"{grupo}.{gmf_codigo}"
                    if clave in STATE.tiempos_gmf:
                        segs = int((ahora - STATE.tiempos_gmf[clave]).total_seconds())
                        STATE.tiempos_congelados[clave] = segs
                        STATE.tiempos_gmf.pop(clave, None)
        except Exception as e:  # pragma: no cover
            logger.error(f"Error cerrando falla {grupo}.{gmf_codigo}: {e}")

    def _on_continuous(self, grupo: str, val_bool: bool, ahora: datetime):
        """Maneja los cambios en el estado del nodo `CONTINUOUS`."""
        prev = STATE.current_continuous.get(grupo)
        STATE.current_continuous[grupo] = val_bool
        STATE.state_cont_known[grupo] = True
        with STATE.lock:
            if grupo not in STATE.acumulado_continuous:
                STATE.acumulado_continuous[grupo] = timedelta(0)
            STATE.ultima_marca_continuous[grupo] = ahora
        if prev is True and val_bool is False:
            STATE.tg_restaurado_enviado[grupo] = False
        if prev is False and val_bool is True:
            segs_off_antes = int(STATE.acumulado_continuous.get(grupo, timedelta()).total_seconds())
            if (segs_off_antes > CFG.telegram_threshold_seconds) and (not STATE.tg_restaurado_enviado.get(grupo, False)):
                TELEGRAM.send(f'✅ Equipo "{grupo}" restablecido. Tiempo acumulado {int(max(0, segs_off_antes))} s')
                if CFG.telegram_restablecido_unico:
                    STATE.tg_restaurado_enviado[grupo] = True
            _exportar_congeladas_de_grupo(grupo, ahora)
            with STATE.lock:
                quedan_activas = any(k.startswith(grupo + ".") for k in STATE.tiempos_gmf)
                quedan_congeladas = any(k.startswith(grupo + ".") for k in STATE.tiempos_congelados)
                if not quedan_activas and not quedan_congeladas:
                    STATE.tg_enviado[grupo] = False

    def _on_gmf_word(self, grupo: str, suffix_hex: str, value: int, ahora: datetime):
        """Maneja los cambios en los códigos GMF, abriendo o cerrando fallas."""
        key = (grupo, suffix_hex)
        prev_value = STATE.last_word_values.get(key, 0)
        STATE.last_word_values[key] = value
        prev_bits = {i for i in range(16) if (prev_value >> i) & 1}
        curr_bits = {i for i in range(16) if (value >> i) & 1}
        for bit_pos in (curr_bits - prev_bits):
            gmf_base = f"{suffix_hex}{bit_pos:X}".upper()
            gmf_id = f"{grupo}.GMF{gmf_base}"
            with STATE.lock:
                if gmf_id not in STATE.tiempos_gmf and gmf_id not in STATE.tiempos_congelados:
                    STATE.tiempos_gmf[gmf_id] = ahora
        for bit_pos in (prev_bits - curr_bits):
            gmf_base = f"{suffix_hex}{bit_pos:X}".upper()
            gmf_id = f"{grupo}.GMF{gmf_base}"
            asyncio.create_task(self._cerrar_falla(grupo, f"GMF{gmf_base}", ahora))

    def datachange_notification(self, node, val, data):
        """Método callback que se activa cuando cambia un valor en el servidor OPC UA."""
        try:
            ahora = datetime.now()
            nid_str = node.nodeid.to_string()
            if not isinstance(nid_str, str) or ";s=" not in nid_str:
                return
            s = nid_str.split(";s=", 1)[1]
            if "." not in s:
                return
            grupo, campo = s.split(".", 1)
            campo = campo.upper()
            if campo == "CONTINUOUS":
                self._on_continuous(grupo, bool(val), ahora)
                return
            if campo.startswith("GMF") and isinstance(val, int):
                self._on_gmf_word(grupo, campo[3:], val, ahora)
        except Exception as e:  # pragma: no cover
            logger.error(f"Handler datachange error: {e}")


# --------------------------------------------------------------------------------------
# Suscripción por grupo
# --------------------------------------------------------------------------------------
# La función `suscribir_grupos` es el punto de entrada para iniciar las suscripciones
# al servidor OPC UA. Por cada grupo de nodos, crea una suscripción, valida los
# nodos GMF y suscribe a los cambios de datos, lo que permite que el `SubHandler`
# reciba notificaciones.

async def suscribir_grupos(client: Client, ncache: NodeCache) -> list:
    """Crea y gestiona las suscripciones de OPC UA para cada grupo de nodos."""
    subs = []
    for grupo in CFG.node_groups:
        try:
            cache_g = ncache.for_group(grupo)
            if not cache_g.get("gmf_validated", False):
                await _validate_gmfs(client, grupo, cache_g)
            try:
                init_val = await cache_g["continuous"].read_value()
                with STATE.lock:
                    STATE.current_continuous[grupo] = bool(init_val)
                    STATE.state_cont_known[grupo] = True
                    STATE.acumulado_continuous.setdefault(grupo, timedelta(0))
                    STATE.ultima_marca_continuous[grupo] = datetime.now()
                logger.info(f"Init CONTINUOUS {grupo} = {bool(init_val)}")
            except Exception as e:  # pragma: no cover
                logger.warning(f"No se pudo leer CONTINUOUS inicial de {grupo}: {e}")
            handler = SubHandler(client, ncache)
            params = ua.CreateSubscriptionParameters()
            params.RequestedPublishingInterval = CFG.sampling_ms
            params.RequestedMaxKeepAliveCount = CFG.sub_keepalive_count
            params.RequestedLifetimeCount = max(CFG.sub_lifetime_count, CFG.sub_keepalive_count * 3)
            params.MaxNotificationsPerPublish = 0
            params.Priority = 0
            try:
                subscription = await client.create_subscription(params, handler)
            except TypeError:
                subscription = await client.create_subscription(CFG.sampling_ms, handler)
            await subscription.subscribe_data_change([cache_g["continuous"], cache_g["total_fault"]], queuesize=CFG.sub_queue_size)
            if cache_g["gmf_nodes"]:
                await subscription.subscribe_data_change(cache_g["gmf_nodes"], queuesize=CFG.sub_queue_size)
            subs.append(subscription)
            logger.info(f"Suscrito grupo {grupo}: {2 + len(cache_g['gmf_nodes'])} ítems")
            STATE.limpiar_error_grupo(grupo, datetime.now())
        except Exception as e:
            msg = f"Error en grupo {grupo}: {e}"
            logger.error(msg)
            STATE.marcar_error_grupo(grupo, msg)
            continue
    return subs


# --------------------------------------------------------------------------------------
# Tareas asíncronas
# --------------------------------------------------------------------------------------
# Estas funciones se ejecutan en segundo plano (`asyncio.create_task`).
# `heartbeat_telegram_y_acumulados` se encarga de:
# 1. Acumular el tiempo de inactividad de los equipos cuando `CONTINUOUS` es `0`.
# 2. Enviar notificaciones de Telegram si la inactividad supera un umbral.

async def heartbeat_telegram_y_acumulados():
    """Tarea asíncrona que actualiza el estado, envía alertas por Telegram
    y marca el monitoreo en la base de datos de Oracle.
    """
    while True:
        ahora = datetime.now()
        with STATE.lock:
            # Lógica para acumular tiempo OFF y enviar notificaciones de Telegram
            for grupo in CFG.node_groups:
                STATE.ultima_marca_continuous.setdefault(grupo, ahora)
                STATE.acumulado_continuous.setdefault(grupo, timedelta())
                if STATE.state_cont_known.get(grupo, False) and (not STATE.current_continuous.get(grupo, True)):
                    delta = ahora - STATE.ultima_marca_continuous[grupo]
                    if delta.total_seconds() > 0:
                        STATE.acumulado_continuous[grupo] += delta
                    STATE.ultima_marca_continuous[grupo] = ahora

                    if (STATE.acumulado_continuous[grupo].total_seconds() > CFG.telegram_threshold_seconds):
                        if not STATE.tg_enviado.get(grupo, False):
                            activos_gmf = [k.split(".")[1] for k in STATE.tiempos_gmf if k.startswith(grupo + ".")]
                            mensaje = _msg_telegram(grupo, int(STATE.acumulado_continuous[grupo].total_seconds()), activos_gmf)
                            TELEGRAM.send(mensaje)
                            STATE.tg_enviado[grupo] = True

            # Lógica para actualizar el tick de monitoreo en Oracle
            if (ahora - STATE.last_mon_update).total_seconds() >= 60:
                ORACLE.tick_monitoring(STATE.session_id, ahora)
                STATE.last_mon_update = ahora

        # Pausa para evitar un uso excesivo de la CPU
        await asyncio.sleep(1)


async def watchdog():
    """Tarea de vigilancia para verificar la conectividad OPC UA.
    Si la conexión se pierde, marca un error de conexión en el estado.
    """
    await asyncio.sleep(5)  # Espera inicial para dar tiempo a la conexión
    while True:
        with STATE.lock:
            if STATE.error_activo:
                await asyncio.sleep(1)
                continue
            
            # Si todas las suscripciones de grupo tienen un error
            todos_con_error = all(STATE.error_grupo_activo.get(g, False) for g in CFG.node_groups)
            if todos_con_error:
                if not STATE.error_activo:
                    msg = "Desconexión total del servidor OPC UA"
                    STATE.error_activo = True
                    STATE.error_inicio = datetime.now()
                    STATE.ultimo_error_msg = msg
                    logger.error(msg)
                    ORACLE.insert_error_conn(STATE.session_id, STATE.error_inicio, datetime.now(), msg)
                    TELEGRAM.send("<b>⚠️ Desconexión total del servidor OPC UA.</b>")
            else:
                if STATE.error_activo:
                    # Se restableció la conexión, limpiar el error
                    fin_error = datetime.now()
                    ORACLE.insert_error_conn(STATE.session_id, STATE.error_inicio, fin_error, "Conexión restablecida")
                    msg = f"<b>✅ Conexión con el servidor OPC UA restablecida.</b> Duración: {_fmt_mmss(int((fin_error - STATE.error_inicio).total_seconds()))}"
                    TELEGRAM.send(msg)
                    STATE.error_activo = False
                    STATE.ultimo_error_msg = ""
                    logger.info("Conexión OPC UA restablecida.")

        await asyncio.sleep(1)


async def main():
    """Función principal del script que inicializa el cliente OPC UA,
    crea las tareas asíncronas y gestiona la reconexión.
    """
    logger.info("Iniciando monitor OPC UA...")
    cargar_descripciones()
    STATE.last_mon_update = datetime.now()
    ORACLE.start_monitoring(STATE.session_id, STATE.last_mon_update)
    
    # Tareas en segundo plano
    asyncio.create_task(heartbeat_telegram_y_acumulados())
    asyncio.create_task(watchdog())

    while True:
        try:
            client = Client(url=CFG.url)
            async with client:
                logger.info(f"Conectado a {CFG.url}")
                ncache = NodeCache(client).build_for_groups(CFG.node_groups)
                subs = await suscribir_grupos(client, ncache)
                
                # Bucle de mantenimiento de la conexión
                while client.is_connected:
                    await asyncio.sleep(1)
                
                # Si la conexión se pierde
                logger.warning("Desconexión del cliente OPC UA, intentando reconectar...")
                with STATE.lock:
                    if not STATE.error_activo:
                        STATE.error_activo = True
                        STATE.error_inicio = datetime.now()
                        STATE.ultimo_error_msg = "Desconexión del cliente"
                        ORACLE.insert_error_conn(STATE.session_id, STATE.error_inicio, datetime.now(), STATE.ultimo_error_msg)
                        TELEGRAM.send("<b>⚠️ El cliente se ha desconectado del servidor OPC UA.</b>")
                for sub in subs:
                    await sub.delete()
        except Exception as e:
            msg = f"Error en la conexión principal: {e}"
            logger.error(msg)
            with STATE.lock:
                if not STATE.error_activo:
                    STATE.error_activo = True
                    STATE.error_inicio = datetime.now()
                    STATE.ultimo_error_msg = msg
                    ORACLE.insert_error_conn(STATE.session_id, STATE.error_inicio, datetime.now(), msg)
                    TELEGRAM.send("<b>⚠️ Error de conexión del cliente con el servidor OPC UA.</b>")
        await asyncio.sleep(5)


# --------------------------------------------------------------------------------------
# Dash UI
# --------------------------------------------------------------------------------------
# Esta sección contiene la lógica para la interfaz de usuario web (`Dash`).
# Se crea un dashboard para mostrar el estado del monitoreo en tiempo real.
# Los `callbacks` de Dash leen el estado global (`STATE`) de forma segura
# utilizando el `threading.Lock`.

app = Dash(__name__)
server = app.server

app.layout = html.Div(
    style={'font-family': 'Arial, sans-serif', 'padding': '20px', 'max-width': '900px', 'margin': 'auto'},
    children=[
        html.H1("Monitor de Fallas OPC UA", style={'text-align': 'center', 'color': '#2C3E50'}),
        html.Hr(),
        
        dcc.Interval(id='interval-component', interval=CFG.ui_interval_ms, n_intervals=0),

        html.Div(id='status-message-container', style={'margin-bottom': '20px'}),
        
        html.Div(
            style={'display': 'flex', 'flex-wrap': 'wrap', 'gap': '20px', 'justify-content': 'center'},
            children=[
                html.Div(id=f'group-info-{g}', style={'border': '1px solid #ddd', 'padding': '15px', 'border-radius': '8px', 'flex': '1 1 45%'})
                for g in CFG.node_groups
            ]
        ),
        
        html.Hr(style={'margin-top': '30px'}),
        html.H2("Fallas Activas", style={'text-align': 'center', 'color': '#2C3E50'}),
        dash_table.DataTable(
            id='falla-activa-table',
            columns=[
                {"name": "Grupo", "id": "grupo"},
                {"name": "Código GMF", "id": "gmf"},
                {"name": "Descripción", "id": "descripcion"},
                {"name": "Inicio de Falla", "id": "inicio"},
                {"name": "Tiempo en Falla", "id": "duracion"}
            ],
            style_table={'overflowX': 'auto'},
            style_header={'backgroundColor': '#2C3E50', 'color': 'white', 'fontWeight': 'bold'},
            style_cell={'textAlign': 'left', 'padding': '10px'},
        )
    ]
)


@app.callback(Output('status-message-container', 'children'),
              [Input('interval-component', 'n_intervals')])
def update_status_message(n):
    """Callback para actualizar el mensaje de estado principal (error de conexión)."""
    with STATE.lock:
        if STATE.error_activo:
            return html.Div(
                f"ERROR: {STATE.ultimo_error_msg}",
                style={'background-color': '#F2D7D5', 'border': '1px solid #E74C3C', 'padding': '10px', 'border-radius': '5px', 'color': '#E74C3C', 'text-align': 'center'}
            )
        else:
            return html.Div(
                "Conexión con el servidor OPC UA estable.",
                style={'background-color': '#D5F5E3', 'border': '1px solid #2ECC71', 'padding': '10px', 'border-radius': '5px', 'color': '#2ECC71', 'text-align': 'center'}
            )


@app.callback([Output(f'group-info-{g}', 'children') for g in CFG.node_groups],
              [Input('interval-component', 'n_intervals')])
def update_group_info(n):
    """Callback para actualizar la información de cada grupo de nodos."""
    ahora = datetime.now()
    outputs = []
    with STATE.lock:
        for grupo in CFG.node_groups:
            estado_grupo = "🟢 OK"
            color = "#2ECC71"
            acum_continuous_str = _fmt_mmss(int(STATE.acumulado_continuous.get(grupo, timedelta()).total_seconds()))
            
            if STATE.error_grupo_activo.get(grupo, False):
                estado_grupo = "🔴 ERROR DE GRUPO"
                color = "#E74C3C"
            elif not STATE.current_continuous.get(grupo, True) and STATE.state_cont_known.get(grupo, False):
                estado_grupo = "🟡 EQUIPO DETENIDO"
                color = "#F1C40F"
            
            outputs.append(html.Div([
                html.H3(grupo, style={'color': color}),
                html.P(f"Estado: {estado_grupo}"),
                html.P(f"Tiempo acumulado en 'OFF': {acum_continuous_str}"),
            ]))
    return outputs


@app.callback(Output('falla-activa-table', 'data'),
              [Input('interval-component', 'n_intervals')])
def update_table(n):
    """Callback para actualizar la tabla de fallas activas."""
    data = []
    ahora = datetime.now()
    with STATE.lock:
        for key, inicio in STATE.tiempos_gmf.items():
            grupo, gmf = key.split(".", 1)
            duracion_td = ahora - inicio
            duracion_str = _fmt_mmss(int(duracion_td.total_seconds()))
            desc = _desc_gmf(grupo, gmf)
            data.append({
                "grupo": grupo,
                "gmf": gmf,
                "descripcion": desc,
                "inicio": inicio.strftime("%Y-%m-%d %H:%M:%S"),
                "duracion": duracion_str
            })
        for key, duracion_s in STATE.tiempos_congelados.items():
            grupo, gmf = key.split(".", 1)
            desc = _desc_gmf(grupo, gmf)
            data.append({
                "grupo": grupo,
                "gmf": gmf,
                "descripcion": desc,
                "inicio": "N/A (Congelado)",
                "duracion": _fmt_mmss(duracion_s)
            })
    return data


@server.route('/health')
def health_check():
    """Endpoint para monitoreo de salud del servidor."""
    with STATE.lock:
        status = "ok" if not STATE.error_activo else "error"
        msg = "Monitoreo operativo." if status == "ok" else STATE.ultimo_error_msg
    return Response(f'{{"status": "{status}", "message": "{msg}"}}', mimetype='application/json')


# --------------------------------------------------------------------------------------
# Ejecución
# --------------------------------------------------------------------------------------
# Este bloque principal decide si ejecutar el servidor web `Waitress` o el servidor de
# desarrollo de Dash, y gestiona la ejecución del bucle asíncrono principal (`main`).

if __name__ == '__main__':
    # Usar Waitress si está disponible para producción
    try:
        from waitress import serve
        logger.info(f"Iniciando servidor Waitress en http://{CFG.host}:{CFG.port}")
        # Se ejecuta el servidor asíncrono en un hilo separado
        threading.Thread(target=lambda: asyncio.run(main()), daemon=True).start()
        serve(app.server, host=CFG.host, port=CFG.port)
    except ImportError:
        logger.warning("Waitress no está instalado. Usando el servidor de desarrollo de Flask (no recomendado para producción).")
        logger.info(f"Iniciando servidor de desarrollo en http://{CFG.host}:{CFG.port}")
        # Se ejecuta el servidor asíncrono en un hilo separado
        threading.Thread(target=lambda: asyncio.run(main()), daemon=True).start()
        app.run_server(host=CFG.host, port=CFG.port, debug=False)