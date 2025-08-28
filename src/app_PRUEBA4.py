# -*- coding: utf-8 -*-
"""
Monitor OPC UA + Dash — TOTAL_START gating (CONVEYOR_SH)
--------------------------------------------------------

Qué cambia:
- TODA la lógica de suscripciones ahora respeta TOTAL_START del grupo maestro (por defecto: "CONVEYOR_SH").
- Si TOTAL_START=0: se fuerza CONTINUOUS=0 y se considera equivalente a GMFxxx=0 (no se listan ni acumulan GMF activas).
- Si TOTAL_START=1: si hay bits GMF activos, se abren fallas NUEVAS desde ese instante (no arrastran las que estaban
  cuando TOTAL_START era 0).
- Se suscribe también a TOTAL_START y se guarda su estado en memoria.
- Se agregó env TOTAL_START_GATING_GROUP (por si querés cambiar el grupo maestro).

Notas:
- Solo se "descarta" la actividad GMF del grupo maestro cuando TOTAL_START=0; el resto de grupos siguen su curso normal.
- CONTINUOUS del grupo maestro se fuerza a 0 mientras TOTAL_START=0 (para reflejar tiempo OFF y notificar por Telegram si corresponde).
"""

from __future__ import annotations

import asyncio
import logging
import os
import sys
import threading
import uuid
from contextlib import contextmanager, suppress
from dataclasses import dataclass, field
from datetime import datetime, time, timedelta
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Optional

import requests
from asyncua import Client, ua
from dash import Dash, dcc, html
from dash import dash_table
from dash.dependencies import Input, Output
from flask import Response
from openpyxl import load_workbook

# ===== DB driver preferido =====
try:
    import cx_Oracle as db  # type: ignore
except ModuleNotFoundError:  # pragma: no cover
    import oracledb as db  # type: ignore

# ===== Excepción opcua segun versión =====
try:
    from asyncua.ua.uaerrors import BadNodeIdUnknown  # type: ignore
except Exception:  # pragma: no cover
    try:
        from asyncua.ua.uaerrors._auto import BadNodeIdUnknown  # type: ignore
    except Exception:  # pragma: no cover
        class BadNodeIdUnknown(Exception):
            pass

# ===== (Opcional) .env =====
try:  # pragma: no cover
    from dotenv import load_dotenv  # type: ignore
except Exception:  # pragma: no cover
    load_dotenv = None  # type: ignore

BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
if load_dotenv:
    load_dotenv(BASE_DIR / ".env")

# --------------------------------------------------------------------------------------
# Config util
# --------------------------------------------------------------------------------------

def env_str(name: str, default: str) -> str:
    return os.getenv(name, default)


def env_bool(name: str, default: bool) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    return v.strip().lower() in {"1", "true", "t", "yes", "y"}


def env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except Exception:
        return default


# --------------------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------------------

@dataclass(frozen=True)
class Config:
    url: str = env_str("OPCUA_URL", "opc.tcp://localhost:52250/freeopcua/server/")
    node_groups: list[str] = field(
        default_factory=lambda: [g.strip() for g in env_str("NODE_GROUPS", "BOMBAS_SH,CONVEYOR_SH").split(",") if g.strip()]
    )
    ns: int = env_int("OPCUA_NS", 2)
    start_hex: int = int(env_str("START_HEX", "0x80"), 16)
    end_hex: int = int(env_str("END_HEX", "0xFF"), 16)

    sampling_ms: int = env_int("SAMPLING_MS", 100)
    sub_queue_size: int = env_int("SUB_QUEUE_SIZE", 1000)
    sub_keepalive_count: int = env_int("SUB_KEEPALIVE_COUNT", 100)
    sub_lifetime_count: int = env_int("SUB_LIFETIME_COUNT", env_int("SUB_KEEPALIVE_COUNT", 100) * 10)

    ui_interval_ms: int = int(float(env_str("INTERVALO_SEGUNDOS", "0.5")) * 1000)

    log_level: str = env_str("LOG_LEVEL", "INFO").upper()
    log_file: str = str((BASE_DIR / "monitor.log").resolve())

    telegram_enabled: bool = env_bool("TELEGRAM_ENABLED", True)
    telegram_token: Optional[str] = os.getenv("TELEGRAM_TOKEN")
    telegram_chat_ids: list[str] = field(
        default_factory=lambda: [x.strip() for x in os.getenv("TELEGRAM_CHAT_IDS", "1600754452,5015132163").split(",") if x.strip()]
    )
    telegram_threshold_seconds: int = env_int("TELEGRAM_THRESHOLD_SECONDS", 60)
    telegram_restablecido_unico: bool = env_bool("TELEGRAM_RESTABLECIDO_UNICO", True)

    host: str = env_str("HOST", "0.0.0.0")
    port: int = env_int("PORT", 8050)

    # Grupo cuyo TOTAL_START actúa como "gate" (por defecto, CONVEYOR_SH)
    gating_group: str = env_str("TOTAL_START_GATING_GROUP", "CONVEYOR_SH").strip()


CFG = Config()

# --------------------------------------------------------------------------------------
# Logging
# --------------------------------------------------------------------------------------

logger = logging.getLogger("opcua-monitor")
logger.setLevel(CFG.log_level)
logger.propagate = False
fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s", "%Y-%m-%d %H:%M:%S")
logger.handlers.clear()
fh = RotatingFileHandler(CFG.log_file, maxBytes=1_000_000, backupCount=3, encoding="utf-8")
fh.setFormatter(fmt)
sh = logging.StreamHandler(sys.stdout)
sh.setFormatter(fmt)
logger.addHandler(fh)
logger.addHandler(sh)

# --------------------------------------------------------------------------------------
# Oracle client
# --------------------------------------------------------------------------------------

def _oracle_conn():
    user = os.getenv("ORACLE_USER")
    pwd = os.getenv("ORACLE_PWD")
    dsn = os.getenv("ORACLE_DSN")
    if not all([user, pwd, dsn]):
        raise RuntimeError("Faltan ORACLE_USER/ORACLE_PWD/ORACLE_DSN en .env")
    return db.connect(user=user, password=pwd, dsn=dsn)


@contextmanager
def oracle_cursor():
    conn = _oracle_conn()
    cur = conn.cursor()
    try:
        yield cur
        conn.commit()
    finally:
        with suppress(Exception):
            cur.close()
            conn.close()


class OracleRepo:
    def ensure_schema(self):
        try:
            with oracle_cursor() as cur:
                cur.execute(
                    """
                    BEGIN
                        EXECUTE IMMEDIATE '
                            CREATE TABLE MONITOREO_OPCUA (
                                SESSION_ID    VARCHAR2(64),
                                TIPO          VARCHAR2(30),
                                GRUPO         VARCHAR2(50),
                                INICIO        TIMESTAMP,
                                FIN           TIMESTAMP,
                                MENSAJE       VARCHAR2(4000)
                            )';
                    EXCEPTION WHEN OTHERS THEN
                        IF SQLCODE != -955 THEN RAISE; END IF;
                    END;"""
                )
                # Intentar agregar EVENT_ID si no existe (ORA-01430)
                with suppress(Exception):
                    cur.execute(
                        """
                        BEGIN
                            EXECUTE IMMEDIATE 'ALTER TABLE MONITOREO_OPCUA ADD (EVENT_ID VARCHAR2(64))';
                        EXCEPTION WHEN OTHERS THEN
                            IF SQLCODE != -1430 THEN RAISE; END IF;
                        END;"""
                    )
        except Exception as e:  # pragma: no cover
            logger.error(f"Error asegurando tablas Oracle: {e}")

    def start_monitoring(self, session_id: str, inicio: datetime):
        with suppress(Exception):
            with oracle_cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
                    VALUES (:sid, 'MONITOREO', NULL, :ini, :fin, :msg)
                    """,
                    dict(sid=session_id, ini=inicio, fin=inicio, msg=None),
                )

    def tick_monitoring(self, session_id: str, fin: datetime):
        with suppress(Exception):
            with oracle_cursor() as cur:
                cur.execute(
                    """UPDATE MONITOREO_OPCUA SET FIN = :fin WHERE SESSION_ID = :sid AND TIPO = 'MONITOREO'""",
                    dict(fin=fin, sid=session_id),
                )

    def insert_error_conn(self, session_id: str, inicio: datetime, fin: datetime, msg: str):
        with suppress(Exception):
            with oracle_cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
                    VALUES (:sid, 'ERROR_CONEXION', NULL, :ini, :fin, :msg)
                    """,
                    dict(sid=session_id, ini=inicio, fin=fin, msg=(msg or "")[:3999]),
                )

    def open_group_error(self, session_id: str, grupo: str, inicio: datetime, msg: str, eid: Optional[str]):
        with suppress(Exception):
            with oracle_cursor() as cur:
                try:
                    cur.execute(
                        """
                        INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE, EVENT_ID)
                        VALUES (:sid, 'ERROR_GRUPO', :grp, :ini, NULL, :msg, :eid)
                        """,
                        dict(sid=session_id, grp=grupo, ini=inicio, msg=(msg or "")[:3999], eid=eid),
                    )
                except Exception:
                    cur.execute(
                        """
                        INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
                        VALUES (:sid, 'ERROR_GRUPO', :grp, :ini, NULL, :msg)
                        """,
                        dict(sid=session_id, grp=grupo, ini=inicio, msg=(msg or "")[:3999]),
                    )

    def close_group_error(self, session_id: str, grupo: str, fin: datetime, eid: Optional[str]):
        with suppress(Exception):
            with oracle_cursor() as cur:
                if eid:
                    cur.execute(
                        """
                        UPDATE MONITOREO_OPCUA SET FIN = :fin
                         WHERE SESSION_ID = :sid AND TIPO = 'ERROR_GRUPO' AND GRUPO = :grp AND EVENT_ID = :eid AND FIN IS NULL
                        """,
                        dict(fin=fin, sid=session_id, grp=grupo, eid=eid),
                    )
                else:
                    cur.execute(
                        """
                        UPDATE MONITOREO_OPCUA SET FIN = :fin
                         WHERE SESSION_ID = :sid AND TIPO = 'ERROR_GRUPO' AND GRUPO = :grp AND FIN IS NULL
                        """,
                        dict(fin=fin, sid=session_id, grp=grupo),
                    )


ORACLE = OracleRepo()
ORACLE.ensure_schema()

# --------------------------------------------------------------------------------------
# Telegram client
# --------------------------------------------------------------------------------------

class TelegramClient:
    def __init__(self, enabled: bool, token: Optional[str], chat_ids: list[str]) -> None:
        self.enabled = enabled and bool(token) and any(cid.strip().lstrip("-").isdigit() for cid in chat_ids)
        self.token = token
        self.chat_ids = [c.strip() for c in chat_ids if c.strip()]
        self.session = requests.Session()

    def send(self, text: str):
        if not self.enabled or not self.token:
            return
        url = f"https://api.telegram.org/bot{self.token}/sendMessage"
        for cid in self.chat_ids:
            payload = dict(chat_id=cid, text=text, parse_mode="HTML", disable_web_page_preview=True)
            try:
                r = self.session.post(url, json=payload, timeout=(3, 8))
                if r.status_code != 200:
                    logger.error(f"Telegram error chat_id={cid}: {r.status_code} {r.text}")
            except Exception as e:  # pragma: no cover
                logger.error(f"Telegram excepción chat_id={cid}: {e}")


TELEGRAM = TelegramClient(CFG.telegram_enabled, CFG.telegram_token, CFG.telegram_chat_ids)

# --------------------------------------------------------------------------------------
# State
# --------------------------------------------------------------------------------------

@dataclass
class State:
    node_groups: list[str]
    session_id: str = field(default_factory=lambda: uuid.uuid4().hex)

    lock: threading.Lock = field(default_factory=threading.Lock)

    tiempos_gmf: dict[str, datetime] = field(default_factory=dict)
    tiempos_congelados: dict[str, int] = field(default_factory=dict)
    acumulado_continuous: dict[str, timedelta] = field(default_factory=dict)
    ultima_marca_continuous: dict[str, datetime] = field(default_factory=dict)
    descripciones_gmf: dict[str, dict[str, str]] = field(default_factory=dict)

    tg_enviado: dict[str, bool] = field(init=False)
    tg_restaurado_enviado: dict[str, bool] = field(init=False)

    last_mon_update: datetime = field(default_factory=lambda: datetime.min)

    error_activo: bool = False
    error_inicio: Optional[datetime] = None
    ultimo_error_msg: str = ""
    error_grupo_activo: dict[str, bool] = field(init=False)
    error_grupo_inicio: dict[str, Optional[datetime]] = field(init=False)
    error_grupo_msg: dict[str, str] = field(init=False)
    error_grupo_event_id: dict[str, Optional[str]] = field(init=False)

    current_continuous: dict[str, bool] = field(init=False)
    state_cont_known: dict[str, bool] = field(init=False)
    last_word_values: dict[tuple[str, str], int] = field(default_factory=dict)

    # NUEVO: TOTAL_START por grupo
    current_total_start: dict[str, bool] = field(init=False)
    state_ts_known: dict[str, bool] = field(init=False)

    def __post_init__(self):
        self.tg_enviado = {g: False for g in self.node_groups}
        self.tg_restaurado_enviado = {g: False for g in self.node_groups}
        self.error_grupo_activo = {g: False for g in self.node_groups}
        self.error_grupo_inicio = {g: None for g in self.node_groups}
        self.error_grupo_msg = {g: "" for g in self.node_groups}
        self.error_grupo_event_id = {g: None for g in self.node_groups}
        self.current_continuous = {g: True for g in self.node_groups}
        self.state_cont_known = {g: False for g in self.node_groups}
        self.current_total_start = {g: True for g in self.node_groups}
        self.state_ts_known = {g: False for g in self.node_groups}

    # ---- helpers ----
    def marcar_error_grupo(self, grupo: str, mensaje: str, inicio: Optional[datetime] = None):
        ahora = inicio or datetime.now()
        with self.lock:
            ya_activo = self.error_grupo_activo.get(grupo, False)
            self.error_grupo_activo[grupo] = True
            if not ya_activo:
                self.error_grupo_inicio[grupo] = ahora
                eid = uuid.uuid4().hex
                self.error_grupo_event_id[grupo] = eid
            else:
                eid = self.error_grupo_event_id.get(grupo)
            self.error_grupo_msg[grupo] = mensaje
        if not ya_activo:
            ORACLE.open_group_error(self.session_id, grupo, ahora, mensaje, eid)
            logger.warning(f"ERROR_GRUPO abierto: {grupo} @ {ahora} | {mensaje} | event_id={eid}")

    def limpiar_error_grupo(self, grupo: str, fin: Optional[datetime] = None):
        ahora = fin or datetime.now()
        with self.lock:
            estaba = self.error_grupo_activo.get(grupo, False)
            eid = self.error_grupo_event_id.get(grupo)
            self.error_grupo_activo[grupo] = False
            self.error_grupo_event_id[grupo] = None
        if estaba:
            ORACLE.close_group_error(self.session_id, grupo, ahora, eid)
            logger.info(f"ERROR_GRUPO cerrado: {grupo} @ {ahora} | event_id={eid}")


STATE = State(CFG.node_groups)

# --------------------------------------------------------------------------------------
# Utilidades
# --------------------------------------------------------------------------------------

def nid(s: str) -> str:
    return f"ns={CFG.ns};s={s}"


def determinar_turno(fecha: datetime) -> str:
    hora = fecha.time()
    if time(6, 5) <= hora <= time(14, 30):
        return "MAÑANA"
    if time(14, 39) <= hora <= time(22, 54):
        return "TARDE"
    if time(23, 13) <= hora <= time(23, 59) or time(0, 0) <= hora <= time(5, 55):
        return "NOCHE"
    return "FUERA DE TURNO"


def _fmt_mmss(segundos: int) -> str:
    m, s = divmod(max(0, int(segundos)), 60)
    return f"{m:02d}:{s:02d}"


def _normalize_code(codigo) -> str:
    if codigo is None:
        return ""
    if isinstance(codigo, (int, float)):
        s = f"{codigo}"
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s.strip().upper()
    return str(codigo).strip().upper()


def _desc_gmf(grupo: str, gmf: str) -> str:
    return STATE.descripciones_gmf.get(grupo, {}).get(gmf.strip().upper(), "Sin descripción")


def _msg_telegram(grupo: str, segundos: int, activos: list[str]) -> str:
    import html as html_escape

    if not activos:
        return f"⚠️ {html_escape.escape(grupo)} en falla, hace {segundos} s."
    if len(activos) == 1:
        g = activos[0]
        d = _desc_gmf(grupo, g)
        return f"⚠️ {html_escape.escape(grupo)} con falla {html_escape.escape(g)}: {html_escape.escape(d)} , hace {segundos} s."
    partes = [f"{html_escape.escape(g)}: {html_escape.escape(_desc_gmf(grupo, g))}" for g in activos]
    listado = " // ".join(partes[:10])  # antes: "; ".join(partes[:10])
    extra = "" if len(partes) <= 10 else f" (+{len(partes)-10} más)"
    return f"⚠️ {html_escape.escape(grupo)} con fallas <b>{listado}{extra}</b> , hace {segundos} s."


# --------------------------------------------------------------------------------------
# Carga de descripciones (Excel)
# --------------------------------------------------------------------------------------

ARCHIVOS_DESCRIPCIONES = {
    "BOMBAS_SH": str((BASE_DIR / "GMFs_de_Equipos" / "GMF_BOMBAS_SH.xlsx").resolve()),
    "CONVEYOR_SH": str((BASE_DIR / "GMFs_de_Equipos" / "GMF_CONVEYOR_SH.xlsx").resolve()),
}


def cargar_descripciones():
    for grupo, ruta in ARCHIVOS_DESCRIPCIONES.items():
        STATE.descripciones_gmf[grupo] = {}
        try:
            wb = load_workbook(ruta, data_only=True)
            hoja = wb.active
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                codigo_raw, descripcion = fila[:2]
                if codigo_raw and descripcion:
                    codigo = _normalize_code(codigo_raw)
                    if not codigo.startswith("GMF"):
                        codigo = "GMF" + codigo
                    STATE.descripciones_gmf[grupo][codigo] = str(descripcion).strip()
            logger.info(f"[Desc] {grupo}: {len(STATE.descripciones_gmf[grupo])} códigos desde {ruta}")
        except Exception as e:  # pragma: no cover
            logger.error(f"Error cargando descripciones {grupo}: {e}")


# --------------------------------------------------------------------------------------
# Exportar a ORACLE (historial de fallas)
# --------------------------------------------------------------------------------------

def exportar_a_oracle(grupo: str, gmf: str, descripcion: str, inicio: datetime, fin: datetime, acumulado: int):
    try:
        tabla = f"HISTORIAL_FALLAS_{grupo.upper()}"
        with oracle_cursor() as cur:
            cur.execute(
                f"""
                BEGIN
                    EXECUTE IMMEDIATE '
                        CREATE TABLE {tabla} (
                            EQUIPO           VARCHAR2(50),
                            FALLA            VARCHAR2(50),
                            DESCRIPCION      VARCHAR2(255),
                            INICIO_FALLA     TIMESTAMP,
                            FIN_FALLA        TIMESTAMP,
                            TIEMPO_OFF_SEG   NUMBER,
                            TURNO            VARCHAR2(20)
                        )';
                EXCEPTION WHEN OTHERS THEN
                    IF SQLCODE != -955 THEN RAISE; END IF;
                END;"""
            )
            cur.execute(
                f"""
                INSERT INTO {tabla} (EQUIPO, FALLA, DESCRIPCION, INICIO_FALLA, FIN_FALLA, TIEMPO_OFF_SEG, TURNO)
                VALUES (:1, :2, :3, :4, :5, :6, :7)
                """,
                (grupo, gmf, descripcion, inicio, fin, acumulado, determinar_turno(inicio)),
            )
        logger.info(f"Exportado a Oracle: {grupo} | {gmf} | {acumulado} seg")
    except Exception as e:  # pragma: no cover
        logger.error(f"Error exportando a Oracle: {e}")


# --------------------------------------------------------------------------------------
# Exportar congeladas al volver CONTINUOUS=1 (con reset)
# --------------------------------------------------------------------------------------

def _exportar_congeladas_de_grupo(grupo: str, ahora: datetime):
    """Exporta a Oracle las fallas "congeladas" de un grupo (cuando CONTINUOUS estuvo en 0)
    y resetea el acumulado OFF del grupo.
    """
    with STATE.lock:
        pendientes = [
            (k.split(".")[1], seg)
            for k, seg in STATE.tiempos_congelados.items()
            if k.startswith(grupo + ".")
        ]
        segundos_off = int(STATE.acumulado_continuous.get(grupo, timedelta()).total_seconds())

    for gmf, seg in pendientes:
        try:
            inicio = ahora - timedelta(seconds=int(seg))
            descripcion = _desc_gmf(grupo, gmf)
            exportar_a_oracle(grupo, gmf, descripcion, inicio, ahora, segundos_off)
            with STATE.lock:
                STATE.tiempos_congelados.pop(f"{grupo}.{gmf}", None)
        except Exception as e:  # pragma: no cover
            logger.error(f"Error exportando congelada {grupo}.{gmf}: {e}")

    # Reset OFF tras exportar todas
    with STATE.lock:
        STATE.acumulado_continuous[grupo] = timedelta(0)

# Alias por compatibilidad
exportar_congeladas_de_grupo = _exportar_congeladas_de_grupo

# --------------------------------------------------------------------------------------
# Node cache
# --------------------------------------------------------------------------------------

class NodeCache:
    def __init__(self, client: Client):
        self.client = client
        self.cache: dict[str, dict] = {}

    def build_for_groups(self, grupos: list[str]):
        for grupo in grupos:
            total_fault = self.client.get_node(nid(f"{grupo}.TOTAL_FAULT"))
            continuous = self.client.get_node(nid(f"{grupo}.CONTINUOUS"))
            total_start = self.client.get_node(nid(f"{grupo}.TOTAL_START"))
            hex_suffixes = [f"{i:X}" for i in range(CFG.start_hex, CFG.end_hex + 1)]
            node_ids = [ua.NodeId.from_string(nid(f"{grupo}.GMF{suf}")) for suf in hex_suffixes]
            gmf_nodes = [self.client.get_node(n) for n in node_ids]
            self.cache[grupo] = {
                "total_fault": total_fault,
                "continuous": continuous,
                "total_start": total_start,
                "gmf_nodes": gmf_nodes,
                "gmf_suffixes": hex_suffixes,
                "gmf_validated": False,
            }
        return self

    def for_group(self, grupo: str) -> dict:
        return self.cache[grupo]


async def _validate_gmfs(client: Client, grupo: str, ncache: dict):
    valid_nodes, valid_suffixes = [], []
    for node, suf in zip(ncache["gmf_nodes"], ncache["gmf_suffixes"]):
        try:
            await node.read_value()
            valid_nodes.append(node)
            valid_suffixes.append(suf)
        except BadNodeIdUnknown:
            logger.debug(f"{grupo}: GMF{suf} no existe, se omite.")
        except Exception:
            logger.debug(f"{grupo}: GMF{suf} no legible por ahora, se omite.")
    ncache["gmf_nodes"], ncache["gmf_suffixes"], ncache["gmf_validated"] = valid_nodes, valid_suffixes, True


# --------------------------------------------------------------------------------------
# Subscription handler
# --------------------------------------------------------------------------------------

class SubHandler:
    def __init__(self, client: Client, ncache: NodeCache):
        self.client = client
        self.ncache = ncache

    # --- helpers ---
    @staticmethod
    def _bits_from_value(v: int) -> set[int]:
        return {i for i in range(16) if ((v >> i) & 1)}

    @staticmethod
    def _clear_gmfs(grupo: str):
        # Eliminar cualquier falla activa o congelada del grupo (sin exportar)
        with STATE.lock:
            for k in list(STATE.tiempos_gmf.keys()):
                if k.startswith(grupo + "."):
                    STATE.tiempos_gmf.pop(k, None)
            for k in list(STATE.tiempos_congelados.keys()):
                if k.startswith(grupo + "."):
                    STATE.tiempos_congelados.pop(k, None)

    async def _snapshot_gmfs_open_if_active(self, grupo: str, ahora: datetime):
        # Lee todos los words GMF del grupo y abre fallas por cada bit=1
        cache_g = self.ncache.for_group(grupo)
        nodes = cache_g["gmf_nodes"]
        if not nodes:
            return
        try:
            # Leer en paralelo
            values = await asyncio.gather(*[n.read_value() for n in nodes], return_exceptions=True)
        except Exception:
            values = []
        for suf, val in zip(cache_g["gmf_suffixes"], values):
            if isinstance(val, Exception) or not isinstance(val, int):
                # Si no se pudo leer, lo dejamos como 0
                val_int = 0
            else:
                val_int = int(val)
            # Actualizar last_word_values para que el próximo cambio parta de este estado
            STATE.last_word_values[(grupo, suf)] = val_int
            for bit in self._bits_from_value(val_int):
                gmf_base = f"{suf}{bit:X}".upper()
                gmf_id = f"{grupo}.GMF{gmf_base}"
                with STATE.lock:
                    if gmf_id not in STATE.tiempos_gmf:
                        STATE.tiempos_gmf[gmf_id] = ahora

    async def _cerrar_falla(self, grupo: str, gmf_codigo: str, ahora: datetime):
        try:
            descripcion = _desc_gmf(grupo, gmf_codigo)
            with STATE.lock:
                inicio_falla = STATE.tiempos_gmf.get(f"{grupo}.{gmf_codigo}")
                acumulado = int(STATE.acumulado_continuous.get(grupo, timedelta()).total_seconds())
                cont_ok = bool(STATE.current_continuous.get(grupo, True)) and STATE.state_cont_known.get(grupo, False)

            if cont_ok:
                exportar_a_oracle(grupo, gmf_codigo, descripcion, inicio_falla, ahora, acumulado)
                with STATE.lock:
                    STATE.acumulado_continuous[grupo] = timedelta(0)

                if (acumulado > CFG.telegram_threshold_seconds) and (not STATE.tg_restaurado_enviado.get(grupo, False)):
                    TELEGRAM.send(f'✅ Equipo "{grupo}" restablecido. Tiempo acumulado {int(max(0, acumulado))} s')
                    if CFG.telegram_restablecido_unico:
                        STATE.tg_restaurado_enviado[grupo] = True

                with STATE.lock:
                    STATE.tiempos_gmf.pop(f"{grupo}.{gmf_codigo}", None)
                    STATE.tiempos_congelados.pop(f"{grupo}.{gmf_codigo}", None)
            else:
                with STATE.lock:
                    clave = f"{grupo}.{gmf_codigo}"
                    if clave in STATE.tiempos_gmf:
                        segs = int((ahora - STATE.tiempos_gmf[clave]).total_seconds())
                        STATE.tiempos_congelados[clave] = segs
                        STATE.tiempos_gmf.pop(clave, None)
        except Exception as e:  # pragma: no cover
            logger.error(f"Error cerrando falla {grupo}.{gmf_codigo}: {e}")

    def _on_continuous(self, grupo: str, val_bool: bool, ahora: datetime):
        prev = STATE.current_continuous.get(grupo)
        STATE.current_continuous[grupo] = val_bool
        STATE.state_cont_known[grupo] = True
        with STATE.lock:
            if grupo not in STATE.acumulado_continuous:
                STATE.acumulado_continuous[grupo] = timedelta(0)
            STATE.ultima_marca_continuous[grupo] = ahora
        if prev is True and val_bool is False:
            STATE.tg_restaurado_enviado[grupo] = False
        if prev is False and val_bool is True:
            segs_off_antes = int(STATE.acumulado_continuous.get(grupo, timedelta()).total_seconds())
            if (segs_off_antes > CFG.telegram_threshold_seconds) and (not STATE.tg_restaurado_enviado.get(grupo, False)):
                TELEGRAM.send(f'✅ Equipo "{grupo}" restablecido. Tiempo acumulado {int(max(0, segs_off_antes))} s')
                if CFG.telegram_restablecido_unico:
                    STATE.tg_restaurado_enviado[grupo] = True
            _exportar_congeladas_de_grupo(grupo, ahora)
            with STATE.lock:
                quedan_activas = any(k.startswith(grupo + ".") for k in STATE.tiempos_gmf)
                quedan_congeladas = any(k.startswith(grupo + ".") for k in STATE.tiempos_congelados)
                if not quedan_activas and not quedan_congeladas:
                    STATE.tg_enviado[grupo] = False

    async def _on_total_start_async(self, grupo: str, val_bool: bool, ahora: datetime):
        # Guardar estado de TOTAL_START
        STATE.current_total_start[grupo] = val_bool
        STATE.state_ts_known[grupo] = True

        # Gate maestro: sólo reaccionamos al TOTAL_START del grupo configurado
        if grupo != CFG.gating_group:
            return

        if not val_bool:
            # TOTAL_START (maestro) = 0 => aplicar a TODOS los grupos:
            #   - Forzar CONTINUOUS=0
            #   - Borrar GMFs activas/congeladas (sin exportar)
            #   - Resetear last_word_values para detectar flancos al volver a 1
            for g in CFG.node_groups:
                self._on_continuous(g, False, ahora)
                self._clear_gmfs(g)
                cache_g = self.ncache.for_group(g)
                for suf in cache_g["gmf_suffixes"]:
                    STATE.last_word_values[(g, suf)] = 0
            logger.info(f"GATE TOTAL_START=0 de {grupo} aplicado a TODOS los grupos: CONTINUOUS=0 y GMF=0 (limpieza)")
        else:
            # TOTAL_START (maestro) = 1 => abrir nuevas fallas segun snapshot actual en TODOS los grupos
            for g in CFG.node_groups:
                await self._snapshot_gmfs_open_if_active(g, ahora)
            logger.info(f"GATE TOTAL_START=1 de {grupo}: snapshot y apertura de fallas en TODOS los grupos desde ahora")

    def _on_gmf_word(self, grupo: str, suffix_hex: str, value: int, ahora: datetime):
        # Si el gate maestro está en 0, ignoramos GMF de cualquier grupo
        if not STATE.current_total_start.get(CFG.gating_group, True):
            STATE.last_word_values[(grupo, suffix_hex)] = 0  # forzar flancos cuando TOTAL_START vuelva a 1
            return

        key = (grupo, suffix_hex)
        prev_value = STATE.last_word_values.get(key, 0)
        curr_value = int(value)
        STATE.last_word_values[key] = curr_value

        prev_bits = {i for i in range(16) if (prev_value >> i) & 1}
        curr_bits = {i for i in range(16) if (curr_value >> i) & 1}

        # Bits que se activan => abrir fallas
        for bit_pos in (curr_bits - prev_bits):
            gmf_base = f"{suffix_hex}{bit_pos:X}".upper()
            gmf_id = f"{grupo}.GMF{gmf_base}"
            with STATE.lock:
                if gmf_id not in STATE.tiempos_gmf and gmf_id not in STATE.tiempos_congelados:
                    STATE.tiempos_gmf[gmf_id] = ahora

        # Bits que se desactivan => cerrar fallas
        for bit_pos in (prev_bits - curr_bits):
            gmf_base = f"{suffix_hex}{bit_pos:X}".upper()
            asyncio.create_task(self._cerrar_falla(grupo, f"GMF{gmf_base}", ahora))


    def datachange_notification(self, node, val, data):
        try:
            ahora = datetime.now()
            nid_str = node.nodeid.to_string()
            if not isinstance(nid_str, str) or ";s=" not in nid_str:
                return
            s = nid_str.split(";s=", 1)[1]
            if "." not in s:
                return
            grupo, campo = s.split(".", 1)
            campo = campo.upper()
            if campo == "CONTINUOUS":
                val_bool = bool(val)
                if not STATE.current_total_start.get(CFG.gating_group, True):
                    val_bool = False  # gate maestro manda
                self._on_continuous(grupo, val_bool, ahora)
                return

            if campo == "TOTAL_START":
                # Manejo asíncrono para poder leer snapshot si hace falta
                asyncio.create_task(self._on_total_start_async(grupo, bool(val), ahora))
                return
            if campo.startswith("GMF") and isinstance(val, int):
                self._on_gmf_word(grupo, campo[3:], int(val), ahora)
        except Exception as e:  # pragma: no cover
            logger.error(f"Handler datachange error: {e}")


# --------------------------------------------------------------------------------------
# Suscripción por grupo
# --------------------------------------------------------------------------------------

async def suscribir_grupos(client: Client, ncache: NodeCache) -> list:
    subs = []
    for grupo in CFG.node_groups:
        try:
            cache_g = ncache.for_group(grupo)
            if not cache_g.get("gmf_validated", False):
                await _validate_gmfs(client, grupo, cache_g)

            handler = SubHandler(client, ncache)

            # Leer valores iniciales de CONTINUOUS y TOTAL_START
            try:
                init_cont = await cache_g["continuous"].read_value()
                with STATE.lock:
                    STATE.current_continuous[grupo] = bool(init_cont)
                    STATE.state_cont_known[grupo] = True
                    STATE.acumulado_continuous.setdefault(grupo, timedelta(0))
                    STATE.ultima_marca_continuous[grupo] = datetime.now()
                logger.info(f"Init CONTINUOUS {grupo} = {bool(init_cont)}")
            except Exception as e:  # pragma: no cover
                logger.warning(f"No se pudo leer CONTINUOUS inicial de {grupo}: {e}")

            try:
                init_ts = await cache_g["total_start"].read_value()
                STATE.current_total_start[grupo] = bool(init_ts)
                STATE.state_ts_known[grupo] = True
                logger.info(f"Init TOTAL_START {grupo} = {bool(init_ts)}")
                # Aplicar gate inicial si el grupo es el maestro
                await handler._on_total_start_async(grupo, bool(init_ts), datetime.now())
            except Exception as e:  # pragma: no cover
                logger.warning(f"No se pudo leer TOTAL_START inicial de {grupo}: {e}")

            # Crear suscripción
            params = ua.CreateSubscriptionParameters()
            params.RequestedPublishingInterval = CFG.sampling_ms
            params.RequestedMaxKeepAliveCount = CFG.sub_keepalive_count
            params.RequestedLifetimeCount = max(CFG.sub_lifetime_count, CFG.sub_keepalive_count * 3)
            params.MaxNotificationsPerPublish = 0
            params.Priority = 0
            try:
                subscription = await client.create_subscription(params, handler)
            except TypeError:
                subscription = await client.create_subscription(CFG.sampling_ms, handler)

            # Suscribir CONTINUOUS, TOTAL_FAULT y TOTAL_START; luego los GMF words
            nodes_base = [cache_g["continuous"], cache_g["total_fault"], cache_g["total_start"]]
            await subscription.subscribe_data_change(nodes_base, queuesize=CFG.sub_queue_size)
            if cache_g["gmf_nodes"]:
                await subscription.subscribe_data_change(cache_g["gmf_nodes"], queuesize=CFG.sub_queue_size)
            subs.append(subscription)
            logger.info(f"Suscrito grupo {grupo}: {3 + len(cache_g['gmf_nodes'])} ítems")
            STATE.limpiar_error_grupo(grupo, datetime.now())
        except Exception as e:
            msg = f"Error en grupo {grupo}: {e}"
            logger.error(msg)
            STATE.marcar_error_grupo(grupo, msg)
            continue
    return subs


# --------------------------------------------------------------------------------------
# Tareas asíncronas
# --------------------------------------------------------------------------------------

async def heartbeat_telegram_y_acumulados():
    while True:
        ahora = datetime.now()
        with STATE.lock:
            for grupo in CFG.node_groups:
                STATE.ultima_marca_continuous.setdefault(grupo, ahora)
                STATE.acumulado_continuous.setdefault(grupo, timedelta())
                if STATE.state_cont_known.get(grupo, False) and (not STATE.current_continuous.get(grupo, True)):
                    delta = ahora - STATE.ultima_marca_continuous[grupo]
                    if delta.total_seconds() > 0:
                        STATE.acumulado_continuous[grupo] += delta
                    STATE.ultima_marca_continuous[grupo] = ahora
                    segundos_acum = int(STATE.acumulado_continuous[grupo].total_seconds())
                    if (segundos_acum >= CFG.telegram_threshold_seconds) and (not STATE.tg_enviado.get(grupo, False)):
                        STATE.tg_enviado[grupo] = True
                        # GMFs activas + GMFs "congeladas" (se desactivaron con CONTINUOUS=0)
                        activos = [gmf_id.split(".")[1] for gmf_id in STATE.tiempos_gmf if gmf_id.startswith(grupo + ".")]
                        congeladas = [gmf_id.split(".")[1] for gmf_id in STATE.tiempos_congelados if gmf_id.startswith(grupo + ".")]

                        # De-dup manteniendo orden (primero activas, luego congeladas)
                        gmfs_para_msg = list(dict.fromkeys(activos + congeladas))

                        TELEGRAM.send(_msg_telegram(grupo, segundos_acum, gmfs_para_msg))

        await asyncio.sleep(0.2)


async def watchdog_por_grupo(client: Client, ncache: NodeCache):
    while True:
        for grupo in CFG.node_groups:
            try:
                cache_g = ncache.for_group(grupo)
                await asyncio.wait_for(cache_g["continuous"].read_value(), timeout=2.0)
                if STATE.error_grupo_activo.get(grupo, False):
                    STATE.limpiar_error_grupo(grupo, datetime.now())
            except Exception as e:
                STATE.marcar_error_grupo(grupo, f"Error en grupo {grupo}: {e}")
        await asyncio.sleep(2.0)


async def tick_oracle():
    while True:
        ahora = datetime.now()
        if (ahora - STATE.last_mon_update).total_seconds() >= 60:
            ORACLE.tick_monitoring(STATE.session_id, ahora)
            STATE.last_mon_update = ahora
        await asyncio.sleep(1.0)


# --------------------------------------------------------------------------------------
# Loop principal
# --------------------------------------------------------------------------------------

async def main():
    cargar_descripciones()
    ORACLE.ensure_schema()
    inicio_sesion = datetime.now()
    ORACLE.start_monitoring(STATE.session_id, inicio_sesion)
    STATE.last_mon_update = inicio_sesion

    while True:
        try:
            async with Client(url=CFG.url) as client:
                logger.info("Conectado al servidor OPC UA")
                if STATE.error_activo:
                    ORACLE.insert_error_conn(STATE.session_id, STATE.error_inicio, datetime.now(), STATE.ultimo_error_msg)  # type: ignore
                    STATE.error_activo = False
                    STATE.error_inicio = None
                    STATE.ultimo_error_msg = ""
                ncache = NodeCache(client).build_for_groups(CFG.node_groups)
                subs = await suscribir_grupos(client, ncache)
                tasks = [
                    asyncio.create_task(heartbeat_telegram_y_acumulados()),
                    asyncio.create_task(tick_oracle()),
                    asyncio.create_task(watchdog_por_grupo(client, ncache)),
                ]
                try:
                    await asyncio.Future()  # dormir "para siempre"
                finally:
                    for t in tasks:
                        t.cancel()
                    for s in subs:
                        with suppress(Exception):
                            await s.delete()
        except Exception as e:  # Reconexión
            if not STATE.error_activo:
                STATE.error_activo = True
                STATE.error_inicio = datetime.now()
                STATE.ultimo_error_msg = str(e)
                logger.error(f"Conexión caída: {STATE.ultimo_error_msg}")
            for grupo in CFG.node_groups:
                STATE.marcar_error_grupo(grupo, f"Conexión caída: {e}", inicio=STATE.error_inicio)
            await asyncio.sleep(2)


def thread_monitor():
    asyncio.run(main())


# --------------------------------------------------------------------------------------
# UI (Dash)
# --------------------------------------------------------------------------------------

BG = "#0f172a"
CARD = "#111827"
TEXT = "#e5e7eb"
MUTED = "#9ca3af"
ACCENT = "#22d3ee"
ACCENT_SOFT = "rgba(34,211,238,0.15)"

container_style = {
    "backgroundColor": BG,
    "minHeight": "100vh",
    "padding": "1px 22px",
    "fontFamily": "Inter, system-ui, -apple-system, Segoe UI, Roboto, Ubuntu, Cantarell, Noto Sans, Helvetica, Arial",
}

title_style = {"color": TEXT, "fontSize": "24px", "marginBottom": "10px", "fontWeight": 700}
card_style = {"backgroundColor": CARD, "borderRadius": "14px", "padding": "14px 16px", "boxShadow": "0 4px 16px rgba(0,0,0,0.35)", "border": f"1px solid {ACCENT_SOFT}"}
table_card_style = {**card_style, "padding": "6px 8px"}

app = Dash(__name__)
app.layout = html.Div(
    [
        dcc.Store(id="store-snapshot", storage_type="local"),
        html.Div(
            [
                html.Div(
                    [
                        html.H3("🛠️ Fallas Activas", style=title_style),
                        html.Div(
                            [
                                dash_table.DataTable(
                                    id="tabla-fallas",
                                    columns=[
                                        {"name": "Equipo", "id": "Equipo"},
                                        {"name": "GMF", "id": "GMF"},
                                        {"name": "Descripción", "id": "Descripcion"},
                                        {"name": "Tiempo (mm:ss)", "id": "TiempoFmt"},
                                        {"name": "Tiempo OFF (seg)", "id": "TiempoOffSeg"},
                                    ],
                                    data=[],
                                    sort_action="native",
                                    page_action="native",
                                    page_size=12,
                                    style_as_list_view=True,
                                    style_table={"height": "78vh", "overflowY": "auto", "border": f"1px solid {ACCENT_SOFT}", "borderRadius": "12px"},
                                    style_header={"backgroundColor": "#111827", "color": TEXT, "fontWeight": "700", "borderBottom": f"2px solid {ACCENT_SOFT}", "position": "sticky", "top": 0, "zIndex": 1},
                                    style_cell={"backgroundColor": "#0b1220", "color": TEXT, "padding": "8px 10px", "border": "0px", "fontSize": "22px", "whiteSpace": "nowrap", "textOverflow": "ellipsis", "maxWidth": 380},
                                    style_data_conditional=[
                                        {"if": {"filter_query": "{TiempoOffSeg} >= 60"}, "backgroundColor": "#783c3c", "color": "#ffdddd", "fontWeight": "700", "borderLeft": "4px solid #ff4d4f"},
                                        {"if": {"filter_query": "{TiempoOffSeg} >= 30 && {TiempoOffSeg} < 60"}, "backgroundColor": "#3f3b1d", "color": "#fff1b8", "fontWeight": "700", "borderLeft": "4px solid #facc15"},
                                        {"if": {"row_index": "odd", "filter_query": "{TiempoOffSeg} < 30"}, "backgroundColor": "#0d1424"},
                                        {"if": {"state": "active"}, "border": f"1px solid {ACCENT_SOFT}"},
                                        {"if": {"column_id": "TiempoFmt"}, "fontWeight": "700", "color": ACCENT},
                                    ],
                                ),
                                html.Div(id="last-update", style={"color": MUTED, "fontSize": "12px", "marginTop": "8px"}),
                            ],
                            style=table_card_style,
                        ),
                    ],
                    style={"flexBasis": "75%", "minWidth": 0},
                ),
                html.Div(
                    [
                        html.H3("🚨 Errores de conexión", style=title_style),
                        html.Div(
                            [
                                dash_table.DataTable(
                                    id="tabla-errores",
                                    columns=[
                                        {"name": "Grupo", "id": "Grupo"},
                                        {"name": "Desde", "id": "Desde"},
                                        {"name": "Hace", "id": "Hace"},
                                        {"name": "Msj", "id": "Mensaje"},
                                    ],
                                    data=[],
                                    sort_action="native",
                                    page_action="native",
                                    page_size=8,
                                    style_as_list_view=True,
                                    style_table={"height": "78vh", "overflowY": "auto", "border": f"1px solid {ACCENT_SOFT}", "borderRadius": "12px"},
                                    style_header={"backgroundColor": "#111827", "color": TEXT, "fontWeight": "700", "borderBottom": f"2px solid {ACCENT_SOFT}", "position": "sticky", "top": 0, "zIndex": 1},
                                    style_cell={"backgroundColor": "#0b1220", "color": TEXT, "padding": "8px 10px", "border": "0px", "fontSize": "18px", "whiteSpace": "nowrap", "textOverflow": "ellipsis", "maxWidth": 600},
                                    style_data_conditional=[
                                        {"if": {"row_index": "odd"}, "backgroundColor": "#0d1424"},
                                        {"if": {"state": "active"}, "border": f"1px solid {ACCENT_SOFT}"},
                                        {"if": {"column_id": "Hace"}, "fontWeight": "700", "color": ACCENT},
                                        {"if": {"filter_query": "{Mensaje} contains 'Error' || {Mensaje} contains 'Failed'"}, "backgroundColor": "#5b1a1a", "color": "#ffe5e5", "borderLeft": "4px solid #ef4444", "fontWeight": "700"},
                                    ],
                                ),
                            ],
                            style=table_card_style,
                        ),
                    ],
                    style={"flexBasis": "25%", "minWidth": 0},
                ),
            ],
            style={"display": "flex", "flexDirection": "row", "gap": "16px", "alignItems": "stretch", "width": "100%"},
        ),
        dcc.Interval(id="intervalo", interval=CFG.ui_interval_ms, n_intervals=0),
    ],
    style=container_style,
)


@app.callback(Output("store-snapshot", "data"), Input("intervalo", "n_intervals"))
def calcular_snapshot(_):
    ahora = datetime.now()
    filas_fallas, filas_errores = [], []
    with STATE.lock:
        for gmf_id, inicio in STATE.tiempos_gmf.items():
            equipo, gmf = gmf_id.split(".")
            segundos = int((ahora - inicio).total_seconds())
            tiempo_off_seg = int(STATE.acumulado_continuous.get(equipo, timedelta()).total_seconds())
            filas_fallas.append({"Equipo": equipo, "GMF": gmf, "Descripcion": _desc_gmf(equipo, gmf), "TiempoFmt": _fmt_mmss(segundos), "TiempoOffSeg": tiempo_off_seg})
        for gmf_id, seg_congelado in STATE.tiempos_congelados.items():
            if gmf_id not in STATE.tiempos_gmf:
                equipo, gmf = gmf_id.split(".")
                tiempo_off_seg = int(STATE.acumulado_continuous.get(equipo, timedelta()).total_seconds())
                filas_fallas.append({"Equipo": equipo, "GMF": gmf, "Descripcion": _desc_gmf(equipo, gmf), "TiempoFmt": _fmt_mmss(seg_congelado), "TiempoOffSeg": tiempo_off_seg})
        for grupo in CFG.node_groups:
            if STATE.error_grupo_activo.get(grupo):
                ini = STATE.error_grupo_inicio.get(grupo) or ahora
                segundos = max(0, int((ahora - ini).total_seconds()))
                filas_errores.append({"Grupo": grupo, "Desde": ini.strftime("%Y-%m-%d %H:%M:%S"), "Hace": _fmt_mmss(segundos), "Mensaje": STATE.error_grupo_msg.get(grupo, "Error de conexión")})
        if STATE.error_activo and STATE.error_inicio:
            segs_g = max(0, int((ahora - STATE.error_inicio).total_seconds()))
            filas_errores.insert(0, {"Grupo": "GLOBAL", "Desde": STATE.error_inicio.strftime("%Y-%m-%d %H:%M:%S"), "Hace": _fmt_mmss(segs_g), "Mensaje": STATE.ultimo_error_msg or "Error de conexión"})
    return {"fallas": filas_fallas, "errores": filas_errores, "last": f"Última actualización: {ahora.strftime('%H:%M:%S')}"}


@app.callback(Output("tabla-fallas", "data"), Output("tabla-errores", "data"), Output("last-update", "children"), Input("store-snapshot", "data"))
def renderizar_tablas(store_data):
    if not store_data:
        return [], [], ""
    return store_data.get("fallas", []), store_data.get("errores", []), store_data.get("last", "")


# --------------------------------------------------------------------------------------
# WSGI / Waitress
# --------------------------------------------------------------------------------------

server = app.server


def start_monitor_once():
    if not getattr(start_monitor_once, "_started", False):
        monitor_thread = threading.Thread(target=thread_monitor, daemon=True)
        monitor_thread.start()
        start_monitor_once._started = True
        logger.info("Monitor thread iniciado (Windows/Waitress)")


start_monitor_once()


@server.route("/healthz")
def healthz():
    return Response("ok", status=200, mimetype="text/plain")


if __name__ == "__main__":
    from waitress import serve  # type: ignore
    logger.info(f"Sirviendo con Waitress en http://{CFG.host}:{CFG.port}")
    serve(server, host=CFG.host, port=CFG.port)
