# ================== IMPORTS ==================
import asyncio
import threading
from datetime import datetime, timedelta, time
import os
import uuid
from typing import Optional
from pathlib import Path
import sys
import logging
from logging.handlers import RotatingFileHandler
from collections import defaultdict

try:
    import cx_Oracle as db
except ModuleNotFoundError:
    import oracledb as db

from asyncua import Client, ua
from dash import Dash, dcc, html
from dash.dependencies import Input, Output
from openpyxl import load_workbook
from dash import dash_table
import requests
import html as html_escape  # para escapar texto sin pisar dash.html
from flask import Response  # Dash usa Flask debajo

# (Opcional) Cargar .env si existiera (una sola vez)
try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover
    load_dotenv = None

# ================== RUTAS Y ENTORNO ==================
# Cuando es .py: base es la carpeta del script
# Cuando es .exe one-file: PyInstaller usa _MEIPASS (carpeta temporal)
BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
if load_dotenv:
    # Carga variables del .env ubicado junto al .py o dentro del bundle
    load_dotenv(BASE_DIR / ".env")

# -------------------- Archivos Excel para descripción (ÚNICO lugar) --------------------
ARCHIVOS_DESCRIPCIONES = {
    "BOMBAS_SH": str((BASE_DIR / "GMFs_de_Equipos" / "GMF_BOMBAS_SH.xlsx").resolve()),
    "CONVEYOR_SH": str((BASE_DIR / "GMFs_de_Equipos" / "GMF_CONVEYOR_SH.xlsx").resolve()),
}

# ================== CONFIGURACIÓN GENERAL ==================
URL = os.getenv("OPCUA_URL", "opc.tcp://localhost:52250/freeopcua/server/")
NODE_GROUPS = os.getenv("NODE_GROUPS", "BOMBAS_SH,CONVEYOR_SH").split(",")
NODE_GROUPS = [g.strip() for g in NODE_GROUPS if g.strip()]
START_HEX = int(os.getenv("START_HEX", "0x80"), 16)
END_HEX = int(os.getenv("END_HEX", "0xFF"), 16)
INTERVALO_SEGUNDOS = float(os.getenv("INTERVALO_SEGUNDOS", "0.05"))  # ahora solo afecta el refresco del Dash

# ================== TELEGRAM ==================
TELEGRAM_ENABLED = os.getenv("TELEGRAM_ENABLED", "true").lower() == "true"
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")  # sin fallback por seguridad
CHAT_IDS_ENV = os.getenv("TELEGRAM_CHAT_IDS")
if CHAT_IDS_ENV:
    TELEGRAM_CHAT_IDS = [x.strip() for x in CHAT_IDS_ENV.split(",") if x.strip()]
else:
    TELEGRAM_CHAT_IDS = [
        "1600754452",
        "5015132163",
    ]

TELEGRAM_THRESHOLD_SECONDS = int(os.getenv("TELEGRAM_THRESHOLD_SECONDS", "60"))  # prod=180
TELEGRAM_RESTABLECIDO_UNICO = os.getenv("TELEGRAM_RESTABLECIDO_UNICO", "true").lower() == "true"

# ================== LOGGING ==================
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
LOG_FILE = os.getenv("LOG_FILE", str((BASE_DIR / "monitor.log").resolve()))
logger = logging.getLogger("opcua-monitor")
logger.setLevel(LOG_LEVEL)

fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s", "%Y-%m-%d %H:%M:%S")
fh = RotatingFileHandler(LOG_FILE, maxBytes=1_000_000, backupCount=3, encoding="utf-8")
fh.setFormatter(fmt)
sh = logging.StreamHandler(sys.stdout)
sh.setFormatter(fmt)
logger.handlers.clear()
logger.addHandler(fh)
logger.addHandler(sh)

# ================== ESTADO ==================
lock = threading.Lock()

tiempos_gmf = {}             # { "GRUPO.GMFxxxx": datetime_inicio }
tiempos_congelados = {}      # { "GRUPO.GMFxxxx": segundos_congelados (si bajó con CONTINUOUS=0) }
acumulado_continuous = {}    # { "GRUPO": timedelta_off }
ultima_marca_continuous = {} # { "GRUPO": datetime_ultima_marca }
descripciones_gmf = {}

# Telegram: flags por grupo
tg_enviado = {g: False for g in NODE_GROUPS}                # aviso "en falla hace X s"
tg_restaurado_enviado = {g: False for g in NODE_GROUPS}     # "restablecido" único por ciclo
last_continuous_state = {g: None for g in NODE_GROUPS}      # seguimiento de transición CONTINUOUS

# Monitoreo Oracle
SESSION_ID = uuid.uuid4().hex
last_mon_update = datetime.min

# Error de conexión (global)
error_activo = False
error_inicio = None
ultimo_error_msg = ""

# Error por grupo (ventanas inicio-fin)
error_grupo_activo = {g: False for g in NODE_GROUPS}
error_grupo_inicio = {g: None for g in NODE_GROUPS}
error_grupo_msg = {g: "" for g in NODE_GROUPS}

# Vida por grupo (para cerrar error al primer evento OK)
last_event_ts = {}  # { "GRUPO": datetime }

# ================== TURNOS ==================
def determinar_turno(fecha: datetime) -> str:
    hora = fecha.time()
    if time(6, 5) <= hora <= time(14, 30):
        return "MAÑANA"
    elif time(14, 39) <= hora <= time(22, 54):
        return "TARDE"
    elif time(23, 13) <= hora <= time(23, 59) or time(0, 0) <= hora <= time(5, 55):
        return "NOCHE"
    return "FUERA DE TURNO"

# ================== ORACLE ==================
def _oracle_conn():
    user = os.getenv("ORACLE_USER")
    pwd  = os.getenv("ORACLE_PWD")
    dsn  = os.getenv("ORACLE_DSN")  # ej: "localhost:1521/orcl"
    if not all([user, pwd, dsn]):
        raise RuntimeError("Faltan ORACLE_USER/ORACLE_PWD/ORACLE_DSN en .env")
    return db.connect(user=user, password=pwd, dsn=dsn)

def _ensure_tablas_oracle():
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        cur.execute(
            """
            BEGIN
                EXECUTE IMMEDIATE '
                    CREATE TABLE MONITOREO_OPCUA (
                        SESSION_ID    VARCHAR2(64),
                        TIPO          VARCHAR2(30),
                        GRUPO         VARCHAR2(50),
                        INICIO        TIMESTAMP,
                        FIN           TIMESTAMP,
                        MENSAJE       VARCHAR2(4000)
                    )';
            EXCEPTION
                WHEN OTHERS THEN
                    IF SQLCODE != -955 THEN RAISE; END IF;
            END;
            """
        )
        conn.commit()
    except Exception as e:
        logger.error(f"Error asegurando tablas Oracle: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def iniciar_monitoreo_en_oracle(inicio: datetime):
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        _ensure_tablas_oracle()
        cur.execute(
            """
            INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
            VALUES (:sid, 'MONITOREO', NULL, :ini, :fin, :msg)
            """,
            dict(sid=SESSION_ID, ini=inicio, fin=inicio, msg=None)
        )
        conn.commit()
        logger.info(f"Sesión de monitoreo iniciada: {inicio} | SESSION_ID={SESSION_ID}")
    except Exception as e:
        logger.error(f"Error iniciando monitoreo en Oracle: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def actualizar_fin_monitoreo(fin: datetime):
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        cur.execute(
            """
            UPDATE MONITOREO_OPCUA
            SET FIN = :fin
            WHERE SESSION_ID = :sid AND TIPO = 'MONITOREO'
            """,
            dict(fin=fin, sid=SESSION_ID)
        )
        conn.commit()
    except Exception as e:
        logger.error(f"Error actualizando fin de monitoreo: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def registrar_error_conexion(inicio: datetime, fin: datetime, mensaje: str):
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        _ensure_tablas_oracle()
        cur.execute(
            """
            INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
            VALUES (:sid, 'ERROR_CONEXION', NULL, :ini, :fin, :msg)
            """,
            dict(sid=SESSION_ID, ini=inicio, fin=fin, msg=(mensaje or "")[:3999])
        )
        conn.commit()
        logger.warning(f"Error de conexión registrado: {inicio} → {fin} | {mensaje}")
    except Exception as e:
        logger.error(f"Error registrando error de conexión: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def abrir_error_grupo(grupo: str, inicio: datetime, mensaje: str):
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        _ensure_tablas_oracle()
        cur.execute(
            """
            INSERT INTO MONITOREO_OPCUA (SESSION_ID, TIPO, GRUPO, INICIO, FIN, MENSAJE)
            VALUES (:sid, 'ERROR_GRUPO', :grp, :ini, NULL, :msg)
            """,
            dict(sid=SESSION_ID, grp=grupo, ini=inicio, msg=(mensaje or "")[:3999])
        )
        conn.commit()
        logger.warning(f"ERROR_GRUPO abierto: {grupo} @ {inicio} | {mensaje}")
    except Exception as e:
        logger.error(f"Error abriendo ERROR_GRUPO ({grupo}): {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

def cerrar_error_grupo(grupo: str, fin: datetime):
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        cur.execute(
            """
            UPDATE MONITOREO_OPCUA
            SET FIN = :fin
            WHERE SESSION_ID = :sid
              AND TIPO = 'ERROR_GRUPO'
              AND GRUPO = :grp
              AND FIN IS NULL
            """,
            dict(fin=fin, sid=SESSION_ID, grp=grupo)
        )
        conn.commit()
        logger.info(f"ERROR_GRUPO cerrado: {grupo} @ {fin}")
    except Exception as e:
        logger.error(f"Error cerrando ERROR_GRUPO ({grupo}): {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

# ================== HELPERS DESCRIPCIONES ==================
def _normalize_code(codigo) -> str:
    if codigo is None:
        return ""
    if isinstance(codigo, (int, float)):
        s = f"{codigo}"
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s.strip().upper()
    return str(codigo).strip().upper()

def _desc_gmf(grupo: str, gmf_codigo: str) -> str:
    return descripciones_gmf.get(grupo, {}).get(gmf_codigo.strip().upper(), "Sin descripción")

def _construir_mensaje_telegram(grupo: str, segundos: int, activos: list[str]) -> str:
    if not activos:
        return f'⚠️ {html_escape.escape(grupo)} en falla hace {segundos} s.'
    if len(activos) == 1:
        g = activos[0]
        d = _desc_gmf(grupo, g)
        return f'⚠️ {html_escape.escape(grupo)} con falla {html_escape.escape(g)}: {html_escape.escape(d)} hace {segundos} s.'
    partes = [f'{html_escape.escape(g)}: {html_escape.escape(_desc_gmf(grupo, g))}' for g in activos]
    listado = "; ".join(partes[:10])
    extra = "" if len(partes) <= 10 else f" (+{len(partes)-10} más)"
    return f'⚠️ {html_escape.escape(grupo)} con fallas <b>{listado}{extra}</b> hace {segundos} s.'

# ================== TELEGRAM ==================
def _telegram_config_ok() -> bool:
    if not TELEGRAM_ENABLED:
        logger.info("Telegram desactivado por configuración")
        return False
    if not TELEGRAM_TOKEN:
        logger.warning("Falta TELEGRAM_TOKEN. Deshabilitando envíos Telegram.")
        return False
    if not TELEGRAM_CHAT_IDS or not any(x.strip() and x.strip().lstrip("-").isdigit() for x in TELEGRAM_CHAT_IDS):
        logger.warning("TELEGRAM_CHAT_IDS vacío o inválido. Deshabilitando Telegram.")
        return False
    return True

def enviar_telegram_a_todos(body: str):
    if not _telegram_config_ok():
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    sess = requests.Session()
    for cid in TELEGRAM_CHAT_IDS:
        cid = cid.strip()
        if not cid or not cid.lstrip("-").isdigit():
            logger.warning(f"chat_id inválido en lista: {cid!r}. Omitiendo.")
            continue
        payload = {
            "chat_id": cid,
            "text": body,
            "parse_mode": "HTML",
            "disable_web_page_preview": True,
        }
        try:
            r = sess.post(url, json=payload, timeout=(3, 8))
            if r.status_code == 200:
                logger.info(f"Telegram enviado a chat_id={cid}")
            else:
                logger.error(f"Error enviando a chat_id={cid}: {r.status_code} {r.text}")
        except Exception as e:
            logger.error(f"Excepción enviando a chat_id={cid}: {e}")

def enviar_telegram_restaurado(grupo: str, segundos: int):
    if not _telegram_config_ok():
        return
    grupo_txt = html_escape.escape(grupo)
    cuerpo = f'✅ Equipo "{grupo_txt}" restablecido. Tiempo acumulado {int(max(0, segundos))} s'
    enviar_telegram_a_todos(cuerpo)

# ================== CARGA DE DESCRIPCIONES ==================
def cargar_descripciones():
    for grupo, ruta in ARCHIVOS_DESCRIPCIONES.items():
        descripciones_gmf[grupo] = {}
        try:
            wb = load_workbook(ruta, data_only=True)
            hoja = wb.active
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                codigo_raw, descripcion = fila[:2]
                if codigo_raw and descripcion:
                    codigo = _normalize_code(codigo_raw)
                    if not codigo.startswith("GMF"):
                        codigo = "GMF" + codigo
                    descripciones_gmf[grupo][codigo] = str(descripcion).strip()
        except Exception as e:
            logger.error(f"Error al cargar descripciones de {grupo} desde {ruta}: {e}")

# ================== EXPORTAR A ORACLE (historial de fallas) ==================
def exportar_a_oracle(grupo, gmf, descripcion, inicio, fin, acumulado):
    try:
        conn = _oracle_conn(); cur = conn.cursor()
        tabla = f"HISTORIAL_FALLAS_{grupo.upper()}"
        turno = determinar_turno(inicio)
        cur.execute(
            f"""
            BEGIN
                EXECUTE IMMEDIATE '
                    CREATE TABLE {tabla} (
                        EQUIPO VARCHAR2(50),
                        FALLA VARCHAR2(50),
                        DESCRIPCION VARCHAR2(255),
                        INICIO_FALLA TIMESTAMP,
                        FIN_FALLA TIMESTAMP,
                        TIEMPO_OFF_SEG NUMBER,
                        TURNO VARCHAR2(20)
                    )';
            EXCEPTION WHEN OTHERS THEN
                IF SQLCODE != -955 THEN RAISE; END IF;
            END;
            """
        )
        cur.execute(
            f"""
            INSERT INTO {tabla} (EQUIPO, FALLA, DESCRIPCION, INICIO_FALLA, FIN_FALLA, TIEMPO_OFF_SEG, TURNO)
            VALUES (:1, :2, :3, :4, :5, :6, :7)
            """,
            (grupo, gmf, descripcion, inicio, fin, acumulado, turno)
        )
        conn.commit()
        logger.info(f"Exportado a Oracle: {grupo} | {gmf} | {turno} | {acumulado} seg")
    except Exception as e:
        logger.error(f"Error al exportar a Oracle: {e}")
    finally:
        try: cur.close(); conn.close()
        except: pass

# ================== SUBSCRIPCIONES (MonitoredItems) ==================
# Eventos entrantes desde OPC UA
event_queue: asyncio.Queue = asyncio.Queue()

# Bancos GMF previos para detectar bits que suben/bajan
prev_bank_value = {}  # { (grupo, suffix_hex): int }

# Exportaciones pendientes cuando GMF termina con CONTINUOUS=0
pending_export = defaultdict(list)  # { grupo: [ {gmf, desc, inicio, fin, acumulado} ] }

# Mapa NodeId -> metadata para enrutar eventos rápido
node_meta = {}  # { nodeid_str: ("cont"|"fault"|"start"|"bank", grupo, suffix?) }

class SubHandler:
    """Recibe notificaciones y las vuelca a una cola asyncio (no bloquear)."""
    def __init__(self, queue: asyncio.Queue):
        self.queue = queue

    def datachange_notification(self, node, val, data):
        try:
            nodeid = node.nodeid.to_string()
            ts = datetime.now()
            self.queue.put_nowait((nodeid, val, ts))
        except Exception as e:
            logger.error(f"Handler datachange error: {e}")

async def build_subscriptions(client: Client, grupos: list[str]):
    """Crea una suscripción y monitorea TOTAL_START, TOTAL_FAULT, CONTINUOUS, y GMF80..FF por grupo."""
    global node_meta
    node_meta.clear()

    nodes = []
    for grupo in grupos:
        total_start = client.get_node(f"ns=2;s={grupo}.TOTAL_START")
        total_fault = client.get_node(f"ns=2;s={grupo}.TOTAL_FAULT")
        continuous  = client.get_node(f"ns=2;s={grupo}.CONTINUOUS")

        for n, kind in [(total_start, "start"), (total_fault, "fault"), (continuous, "cont")]:
            node_meta[n.nodeid.to_string()] = (kind, grupo)
            nodes.append(n)

        # Bancos GMF: 0x80..0xFF (enteros 16-bit)
        for i in range(START_HEX, END_HEX + 1):
            suffix = f"{i:X}"
            node   = client.get_node(f"ns=2;s={grupo}.GMF{suffix}")
            node_meta[node.nodeid.to_string()] = ("bank", grupo, suffix)
            nodes.append(node)

    handler = SubHandler(event_queue)
    # publish_interval=100 ms (ajustable)
    sub = await client.create_subscription(100, handler)
    await sub.subscribe_data_change(nodes)
    return sub

def _gmf_code(suffix_hex: str, bit_pos: int) -> str:
    # bit_pos 0..15 -> 0..F en hex
    return f"GMF{suffix_hex}{bit_pos:X}"

async def process_events():
    """Toma eventos de la cola y actualiza los estados (tiempos, pending_export, etc.). Maneja errores por grupo."""
    while True:
        nodeid, val, ts = await event_queue.get()
        meta = node_meta.get(nodeid)
        if not meta:
            continue

        kind = meta[0]
        grupo = meta[1]  # todos los kinds tienen grupo

        try:
            # ---- marca de vida y cierre de error si estaba abierto ----
            with lock:
                last_event_ts[grupo] = ts
            if error_grupo_activo.get(grupo):
                try:
                    cerrar_error_grupo(grupo, ts)
                except Exception:
                    pass
                with lock:
                    error_grupo_activo[grupo] = False
                    error_grupo_inicio[grupo] = None
                    error_grupo_msg[grupo] = ""

            # ================== LÓGICA NORMAL ==================
            if kind == "cont":
                new = bool(val)
                prev = last_continuous_state.get(grupo)
                last_continuous_state[grupo] = new

                with lock:
                    if grupo not in acumulado_continuous:
                        acumulado_continuous[grupo] = timedelta(0)
                        ultima_marca_continuous[grupo] = ts
                    if new is False:
                        tg_restaurado_enviado[grupo] = False
                        ultima_marca_continuous[grupo] = ts

                if new is True:
                    # Exportar pendientes y limpiar de la vista
                    pend = pending_export.get(grupo, [])
                    if pend:
                        for rec in pend:
                            exportar_a_oracle(grupo, rec["gmf"], rec["desc"], rec["inicio"], rec["fin"], rec["acumulado"])
                            gmf_id = f"{grupo}.{rec['gmf']}"
                            with lock:
                                tiempos_gmf.pop(gmf_id, None)
                                tiempos_congelados.pop(gmf_id, None)
                        pending_export[grupo].clear()

                    # Telegram "restablecido"
                    acum = int(acumulado_continuous.get(grupo, timedelta()).total_seconds())
                    if acum > TELEGRAM_THRESHOLD_SECONDS and not tg_restaurado_enviado.get(grupo, False):
                        asyncio.create_task(asyncio.to_thread(enviar_telegram_restaurado, grupo, acum))
                        if TELEGRAM_RESTABLECIDO_UNICO:
                            tg_restaurado_enviado[grupo] = True

                    # Reset del acumulador si no quedan GMFs activas
                    with lock:
                        if not any(k.startswith(grupo + ".") for k in tiempos_gmf):
                            acumulado_continuous[grupo] = timedelta(0)
                            tg_enviado[grupo] = False

            elif kind == "bank":
                suffix = meta[2]
                if not isinstance(val, int):
                    raise ValueError(f"Valor no entero en banco {grupo}.GMF{suffix}: {val!r}")

                key = (grupo, suffix)
                prev = prev_bank_value.get(key, 0)
                cur  = val & 0xFFFF

                # Bits que suben (0->1)
                raised = [b for b in range(16) if ((prev >> b) & 1) == 0 and ((cur >> b) & 1) == 1]
                # Bits que bajan (1->0)
                lowered = [b for b in range(16) if ((prev >> b) & 1) == 1 and ((cur >> b) & 1) == 0]

                now = ts

                # ALTAS
                for b in raised:
                    gmf = _gmf_code(suffix, b)
                    gmf_id = f"{grupo}.{gmf}"
                    with lock:
                        if gmf_id not in tiempos_gmf:
                            tiempos_gmf[gmf_id] = now
                        # si estaba congelada, ya no: vuelve a contar en vivo
                        tiempos_congelados.pop(gmf_id, None)

                # BAJAS
                for b in lowered:
                    gmf = _gmf_code(suffix, b)
                    gmf_id = f"{grupo}.{gmf}"

                    with lock:
                        inicio = tiempos_gmf.get(gmf_id)  # NO la sacamos aún si CONTINUOUS=0
                    if not inicio:
                        continue

                    desc = _desc_gmf(grupo, gmf)
                    cont_now = bool(last_continuous_state.get(grupo, True))
                    acum = int(acumulado_continuous.get(grupo, timedelta()).total_seconds())

                    if cont_now:
                        # CONTINUOUS=1 → exporto y retiro de la vista
                        exportar_a_oracle(grupo, gmf, desc, inicio, now, acum)
                        with lock:
                            tiempos_gmf.pop(gmf_id, None)
                            tiempos_congelados.pop(gmf_id, None)
                            if not any(k.startswith(grupo + ".") for k in tiempos_gmf):
                                acumulado_continuous[grupo] = timedelta(0)
                                tg_enviado[grupo] = False
                    else:
                        # CONTINUOUS=0 → congelar (sigue visible con tiempo fijo) y exportar luego
                        elapsed = int((now - inicio).total_seconds())
                        with lock:
                            tiempos_congelados[gmf_id] = elapsed
                        pending_export[grupo].append({
                            "gmf": gmf, "desc": desc, "inicio": inicio, "fin": now, "acumulado": acum
                        })

                prev_bank_value[key] = cur

            elif kind == "start":
                pass
            elif kind == "fault":
                pass

        except Exception as e:
            # ---- ERROR POR GRUPO: abrir ventana si no estaba abierta ----
            msg = f"Error procesando evento ({kind}) en {grupo}: {e}"
            logger.error(msg)
            if not error_grupo_activo.get(grupo):
                try:
                    abrir_error_grupo(grupo, ts, msg)
                except Exception:
                    pass
                with lock:
                    error_grupo_activo[grupo] = True
                    error_grupo_inicio[grupo] = ts
                    error_grupo_msg[grupo] = msg
            continue

async def heartbeat():
    """Cada 1 s: acumula downtime si CONTINUOUS=0, dispara Telegram por umbral y actualiza FIN de sesión."""
    global last_mon_update
    while True:
        ahora = datetime.now()

        for grupo in NODE_GROUPS:
            cont = last_continuous_state.get(grupo, True)
            with lock:
                if grupo not in acumulado_continuous:
                    acumulado_continuous[grupo] = timedelta(0)
                    ultima_marca_continuous[grupo] = ahora

                if cont is False:
                    delta = ahora - ultima_marca_continuous[grupo]
                    if delta.total_seconds() > 0:
                        acumulado_continuous[grupo] += delta
                # update marca
                ultima_marca_continuous[grupo] = ahora

                # Telegram “en falla hace X s”
                segundos_acum = int(acumulado_continuous[grupo].total_seconds())
                if (segundos_acum >= TELEGRAM_THRESHOLD_SECONDS) and (not tg_enviado.get(grupo, False)):
                    activos_snapshot = [
                        gmf_id.split(".")[1]
                        for gmf_id in tiempos_gmf.keys()
                        if gmf_id.startswith(grupo + ".")
                    ]
                    tg_enviado[grupo] = True
                    body = _construir_mensaje_telegram(grupo, segundos_acum, activos_snapshot)
                    asyncio.create_task(asyncio.to_thread(enviar_telegram_a_todos, body))

        # Touch de fin de monitoreo cada 60 s
        if (ahora - last_mon_update).total_seconds() >= 60:
            try:
                actualizar_fin_monitoreo(ahora)
            except Exception:
                pass
            last_mon_update = ahora

        await asyncio.sleep(1.0)

# ================== LOOP PRINCIPAL (con suscripciones) ==================
async def main():
    cargar_descripciones()
    # Intentar asegurar tablas al inicio (si no hay Oracle, solo loguea)
    try:
        _ensure_tablas_oracle()
    except Exception:
        pass

    inicio_sesion = datetime.now()
    try:
        iniciar_monitoreo_en_oracle(inicio_sesion)
    except Exception:
        pass

    global error_activo, error_inicio, ultimo_error_msg, last_mon_update
    last_mon_update = inicio_sesion

    # Tareas auxiliares (no se cancelan en reconexión)
    hb_task = asyncio.create_task(heartbeat())
    proc_task = asyncio.create_task(process_events())

    while True:
        sub = None
        try:
            async with Client(url=URL) as client:
                logger.info("Conectado al servidor OPC UA")

                # Si veníamos de error, cerrar ventanas por grupo y registrar error global
                if error_activo:
                    try:
                        registrar_error_conexion(error_inicio, datetime.now(), ultimo_error_msg)
                    except Exception:
                        pass
                    for grupo in NODE_GROUPS:
                        if error_grupo_activo.get(grupo):
                            try:
                                cerrar_error_grupo(grupo, datetime.now())
                            except Exception:
                                pass
                            with lock:
                                error_grupo_activo[grupo] = False
                                error_grupo_inicio[grupo] = None
                                error_grupo_msg[grupo] = ""
                    error_activo = False
                    error_inicio = None
                    ultimo_error_msg = ""

                # Reset básico por reconexión
                prev_bank_value.clear()

                # Crear suscripciones
                sub = await build_subscriptions(client, NODE_GROUPS)

                # Mantener viva la conexión; eventos llegan por handler
                while True:
                    await asyncio.sleep(3600)

        except Exception as e:
            # Abrir errores por grupo la primera vez
            if not error_activo:
                error_activo = True
                error_inicio = datetime.now()
                ultimo_error_msg = str(e)
                logger.error(f"Conexión caída: {ultimo_error_msg}")
                for grupo in NODE_GROUPS:
                    if not error_grupo_activo.get(grupo):
                        inicio = datetime.now()
                        msg = f"Error en grupo {grupo}: {ultimo_error_msg}"
                        try:
                            abrir_error_grupo(grupo, inicio, msg)
                        except Exception:
                            pass
                        with lock:
                            error_grupo_activo[grupo] = True
                            error_grupo_inicio[grupo] = inicio
                            error_grupo_msg[grupo] = msg
            await asyncio.sleep(2)  # backoff corto y reintenta
        finally:
            # Intentamos limpiar la suscripción si existe
            try:
                if sub is not None:
                    await sub.delete()
            except Exception:
                pass

# ================== HILO DE MONITOREO ==================
def thread_monitor():
    asyncio.run(main())

# ================== ESTILOS ==================
BG = "#0f172a"; CARD = "#111827"; TEXT = "#e5e7eb"; MUTED = "#9ca3af"
ACCENT = "#22d3ee"; ACCENT_SOFT = "rgba(34,211,238,0.15)"

container_style = {
    "backgroundColor": BG, "minHeight": "100vh", "padding": "1px 22px",
    "fontFamily": "Inter, system-ui, -apple-system, Segoe UI, Roboto, Ubuntu, Cantarell, Noto Sans, Helvetica, Arial"
}

title_style = {"color": TEXT, "fontSize": "24px", "marginBottom": "10px", "fontWeight": 700}
card_style = {"backgroundColor": CARD, "borderRadius": "14px", "padding": "14px 16px",
              "boxShadow": "0 4px 16px rgba(0,0,0,0.35)", "border": f"1px solid {ACCENT_SOFT}"}

table_card_style = {**card_style, "padding": "6px 8px"}

# ================== DASHBOARD (Dash) ==================
app = Dash(__name__)
app.layout = html.Div([
    html.H3("🛠️ Fallas Activas", style=title_style),

    html.Div([
        dash_table.DataTable(
            id="tabla-fallas",
            columns=[
                {"name": "Equipo", "id": "Equipo"},
                {"name": "GMF", "id": "GMF"},
                {"name": "Descripción", "id": "Descripcion"},
                {"name": "Tiempo (mm:ss)", "id": "TiempoFmt"},
                {"name": "Tiempo OFF (seg)", "id": "TiempoOffSeg"},
            ],
            data=[],
            sort_action="native",
            page_action="native",
            page_size=12,
            style_as_list_view=True,
            style_table={"height": "45vh", "overflowY": "auto",
                         "border": f"1px solid {ACCENT_SOFT}", "borderRadius": "12px"},
            style_header={"backgroundColor": "#111827", "color": TEXT, "fontWeight": "700",
                          "borderBottom": f"2px solid {ACCENT_SOFT}", "position": "sticky", "top": 0, "zIndex": 1},
            style_cell={"backgroundColor": "#0b1220", "color": TEXT, "padding": "8px 10px",
                        "border": "0px", "fontSize": "24px", "whiteSpace": "nowrap",
                        "textOverflow": "ellipsis", "maxWidth": 380},
            style_data_conditional=[
                {"if": {"filter_query": "{TiempoOffSeg} >= 60"},
                 "backgroundColor": "#783c3c", "color": "#ffdddd", "fontWeight": "700",
                 "borderLeft": "4px solid #ff4d4f"},
                {"if": {"filter_query": "{TiempoOffSeg} >= 30 && {TiempoOffSeg} < 60"},
                 "backgroundColor": "#3f3b1d", "color": "#fff1b8", "fontWeight": "700",
                 "borderLeft": "4px solid #facc15"},
                {"if": {"row_index": "odd", "filter_query": "{TiempoOffSeg} < 30"},
                 "backgroundColor": "#0d1424"},
                {"if": {"state": "active"}, "border": f"1px solid {ACCENT_SOFT}"},
                {"if": {"column_id": "TiempoFmt"}, "fontWeight": "700", "color": ACCENT},
            ],
        ),
        html.Div(id="last-update", style={"color": MUTED, "fontSize": "12px", "marginTop": "8px"}),
    ], style=table_card_style),

    html.H3("🚨 Errores de conexión por equipo", style={**title_style, "marginTop": "16px"}),

    html.Div([
        dash_table.DataTable(
            id="tabla-errores",
            columns=[
                {"name": "Grupo", "id": "Grupo"},
                {"name": "Desde", "id": "Desde"},
                {"name": "Hace (mm:ss)", "id": "Hace"},
                {"name": "Mensaje", "id": "Mensaje"},
            ],
            data=[],
            sort_action="native",
            page_action="native",
            page_size=8,
            style_as_list_view=True,
            style_table={"height": "30vh", "overflowY": "auto",
                         "border": f"1px solid {ACCENT_SOFT}", "borderRadius": "12px"},
            style_header={"backgroundColor": "#111827", "color": TEXT, "fontWeight": "700",
                          "borderBottom": f"2px solid {ACCENT_SOFT}", "position": "sticky", "top": 0, "zIndex": 1},
            style_cell={"backgroundColor": "#0b1220", "color": TEXT, "padding": "8px 10px",
                        "border": "0px", "fontSize": "18px", "whiteSpace": "nowrap",
                        "textOverflow": "ellipsis", "maxWidth": 600},
            style_data_conditional=[
                {"if": {"row_index": "odd"}, "backgroundColor": "#0d1424"},
                {"if": {"state": "active"}, "border": f"1px solid {ACCENT_SOFT}"},
                {"if": {"column_id": "Hace"}, "fontWeight": "700", "color": ACCENT},
                {"if": {"filter_query": "{Mensaje} contains 'Error' || {Mensaje} contains 'Failed'"},
                 "backgroundColor": "#5b1a1a", "color": "#ffe5e5", "borderLeft": "4px solid #ef4444", "fontWeight": "700"},
            ],
        ),
    ], style=table_card_style),

    dcc.Interval(id="intervalo", interval=1000, n_intervals=0),
], style=container_style)

# ================== HELPERS & CALLBACKS ==================
def _fmt_mmss(segundos: int) -> str:
    m, s = divmod(max(0, int(segundos)), 60)
    return f"{m:02d}:{s:02d}"

@app.callback(
    Output("tabla-fallas", "data"),
    Output("tabla-errores", "data"),
    Output("last-update", "children"),
    Input("intervalo", "n_intervals"),
)
def actualizar_tablas(_):
    ahora = datetime.now()
    filas_fallas = []
    filas_errores = []

    with lock:
        for gmf_id, inicio in tiempos_gmf.items():
            equipo, gmf = gmf_id.split(".")
            descripcion = _desc_gmf(equipo, gmf)

            # Si está "congelada", mostramos su tiempo fijo; si no, tiempo en vivo
            if gmf_id in tiempos_congelados:
                segundos = int(tiempos_congelados[gmf_id])
            else:
                segundos = int((ahora - inicio).total_seconds())

            tiempo_off_seg = int(acumulado_continuous.get(equipo, timedelta()).total_seconds())
            filas_fallas.append({
                "Equipo": equipo,
                "GMF": gmf,
                "Descripcion": descripcion,
                "TiempoFmt": _fmt_mmss(segundos),
                "TiempoOffSeg": tiempo_off_seg,
            })

        for grupo in NODE_GROUPS:
            if error_grupo_activo.get(grupo):
                ini = error_grupo_inicio.get(grupo) or ahora
                segundos = max(0, int((ahora - ini).total_seconds()))
                filas_errores.append({
                    "Grupo": grupo,
                    "Desde": ini.strftime("%Y-%m-%d %H:%M:%S"),
                    "Hace": _fmt_mmss(segundos),
                    "Mensaje": error_grupo_msg.get(grupo, "Error de conexión"),
                })

    last = f"Última actualización: {ahora.strftime('%H:%M:%S')}"
    return filas_fallas, filas_errores, last

# ================== PRODUCCIÓN WSGI (Waitress) ==================
server = app.server

# --- Arranque único del monitor (evita duplicados) ---
def start_monitor_once():
    if not getattr(start_monitor_once, "_started", False):
        monitor_thread = threading.Thread(target=thread_monitor, daemon=True)
        monitor_thread.start()
        start_monitor_once._started = True
        logger.info("Monitor thread iniciado (Windows/Waitress)")

# Llamar al cargar el módulo (cuando lo importa Waitress/WSGI)
start_monitor_once()

# --- Healthcheck ---
@server.route("/healthz")
def healthz():
    return Response("ok", status=200, mimetype="text/plain")

# --- Entradas de ejecución ---
if __name__ == "__main__":
    # Ejecutar como script: producción con Waitress (sin warning)
    from waitress import serve
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "8050"))
    logger.info(f"Sirviendo con Waitress en http://{host}:{port}")
    serve(server, host=host, port=port)
